import os
import re
import sys
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def diff_memory(dut_fog_folder_path, ref_fog_folder_path):
    
    # INT_RE = re.compile(r'^\s*([^\s:]+)\s*:?\s*([+-]?\d+)\b')
    #keys_to_check = ["MemTotal", "MemFree"]
    
    def parse_mem_file(file_path, get_first_value=True):
        data = OrderedDict()
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Regex to capture key and value, with optional kB unit
                    pattern = r'^\s*([^\s:]+)\s*:?\s*([+-]?\d+)(?:\s*kB)?\s*.*$'

                    m = re.match(pattern, line)
                    if not m:
                        continue
                    key = m.group(1).rstrip(':')
                    try:
                        val = int(m.group(2))
                        # convert kB to MB
                        if 'kB' in line:
                            val /= 1000.0  # convert to MB
                    except ValueError:
                        continue
                    
                    if get_first_value:
                        # keep first occurrence in file if duplicated
                        if key not in data:
                            data[key] = val  # get first value
                    else:
                        # keep last occurrence (re-entry value)
                        data[key] = val  # get re-entry value
                        
        except Exception as e:
            # skip unreadable files but print notice
            print(f"[WARN] cannot read file {file_path}: {e}")
        return data

    def collect_folder_data(folder_path, file_pattern, get_first_value=True):
        app_files = defaultdict(list)
        for fname in sorted(os.listdir(folder_path)):
        # for fname in os.listdir(folder_path):

            #if file_pattern not in fname:
            if not re.search(file_pattern, fname, re.IGNORECASE):
                continue
            fpath = os.path.join(folder_path, fname)
            if os.path.isdir(fpath):
                continue
            try:
                app_name = re.split(file_pattern, fname, flags=re.IGNORECASE)[0].split("_")[-1]
            except Exception:
                app_name = "UNKNOWN"
            app_files[app_name].append((fname, fpath))

        results = {}
        for app, filetuples in app_files.items():
            # filetuples is sorted by filename because we iterated sorted(os.listdir)
            basenames = [t[0] for t in filetuples]
            paths = [t[1] for t in filetuples]

            all_keys = []
            values = dict()  # key -> list

            for idx, p in enumerate(paths):
                parsed = parse_mem_file(p, get_first_value)  # OrderedDict
                # if new key appears in this file, create and pad with None for prior files
                for k in parsed.keys():
                    if k not in all_keys:
                        all_keys.append(k)
                        values[k] = [None] * idx  # pad for previous files

                # append values for all keys in the canonical order
                for k in all_keys:
                    values[k].append(parsed.get(k))  # parsed.get(k) may be None if missing in this file

            # ensure every key has same length list (equal to number of files)
            n_files = len(paths)
            for k in all_keys:
                if len(values[k]) < n_files:
                    values[k].extend([None] * (n_files - len(values[k])))

            results[app] = {
                "files": basenames,
                "paths": paths,
                "keys": all_keys,
                "values": values
            }

        return results

    

    # def get_prefix_from_folder(folder_path):
    #     for fname in sorted(os.listdir(folder_path)):
    #         #if "_Start_" in fname:
    #         if re.search("_start_", fname, re.IGNORECASE):
    #             return fname.split("_")[0]
    #     return "UNKNOWN"

    def get_prefix_from_folder(folder_path):
        for fname in sorted(os.listdir(folder_path)):
            if re.search("_start_", fname, re.IGNORECASE):
                # Separate by "start"
                parts = re.split(r"_start_|_Start_|-start_|-Start_", fname, flags=re.IGNORECASE)
                if len(parts) > 0:
                    # Get string before "start"
                    prefix_part = parts[0]
                    # Split by "-" or "_"
                    prefix_parts = re.split(r"_|-", prefix_part)
                    # Merge to prefix
                    if len(prefix_parts) >= 2:
                        return f"{prefix_parts[0]}_{prefix_parts[1]}" if "_" in prefix_part else f"{prefix_parts[0]}-{prefix_parts[1]}"
                    else:
                        return prefix_part
        return "UNKNOWN"

    def _format_number_for_cell(v):
        if v is None:
            return None
        # if v is already int -> return as int
        if isinstance(v, int):
            return v
        try:
            fv = float(v)
        except Exception:
            return v
        if abs(fv - int(fv)) < 1e-9:
            return int(round(fv))
        # round to 2 decimals
        return round(fv, 2)

    def write_app_sheet(ws, app, d1, d2, prefix1, prefix2):
        """
        Write sheet for a single app.
        Row1: merged cells with prefix1 and prefix2 above their columns
        Row2: Cycle headers + Avg columns + final Diff
        Row3..: keys and values
        """

        files1 = d1["files"] if d1 else []
        files2 = d2["files"] if d2 else []
        n1 = len(files1)
        n2 = len(files2)

        ws.cell(1, 1).value = "Unit: MB"
        ws.cell(1, 1).font = Font(bold=True)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Make red background for average columns 
        red_fill = PatternFill(start_color="e80707", end_color="e80707", fill_type="solid")
        # column mapping:
        # col 1 = Key
        # cols 2..(2+n1-1) = folder1 cycles (if n1>0)
        # col (2+n1) = Avg_F1
        # cols after that = folder2 cycles, then Avg_F2, then Diff
        col = 1
        # Row 1: model names (we will merge ranges)
        # Determine ranges
        start_f1 = 2
        end_f1 = start_f1 + n1  # includes Avg column
        start_f2 = end_f1 + 1
        end_f2 = start_f2 + n2  # includes Avg column
        diff_col = end_f2 + 1

        # If no files in a folder, we still reserve 1 column for Avg.
        if n1 == 0:
            start_f1 = 2
            end_f1 = start_f1  # only avg at col 2
            start_f2 = end_f1 + 1
            end_f2 = start_f2 + n2
            diff_col = end_f2 + 1
        if n2 == 0:
            # recompute in case n2==0
            start_f1 = 2
            end_f1 = start_f1 + n1
            start_f2 = end_f1 + 1
            end_f2 = start_f2  # only avg at start_f2
            diff_col = end_f2 + 1

        # Merge and write model names on row 1
        # Folder1 group is from start_f1 to (start_f1 + n1) inclusive (cycles + avg)
        # But ensure start<=end
        try:
            if start_f1 <= end_f1:
                ws.merge_cells(start_row=1, start_column=start_f1, end_row=1, end_column=end_f1)
                ws.cell(row=1, column=start_f1).value = prefix1
                ws.cell(row=1, column=start_f1).alignment = Alignment(horizontal='center')
                ws.cell(row=1, column=start_f1).font = Font(bold=True)
            if start_f2 <= end_f2:
                ws.merge_cells(start_row=1, start_column=start_f2, end_row=1, end_column=end_f2)
                ws.cell(row=1, column=start_f2).value = prefix2
                ws.cell(row=1, column=start_f2).alignment = Alignment(horizontal='center')
                ws.cell(row=1, column=start_f2).font = Font(bold=True)
        except Exception:
            # ignore merge issues for very small sheets
            pass

        # Row 2: cycles and Avg labels
        # ws.cell(row=2, column=1).value = ""  # Key column header row2 empty
        # Folder1 cycles
        cur = start_f1
        for i in range(n1):
            ws.cell(row=2, column=cur + i).value = f"Cycle{i+1}"
            ws.cell(row=2, column=cur + i).font = Font(bold=True)
            ws.cell(row=2, column=cur + i).alignment = Alignment(horizontal='center')  # center align cycle headers
        avg_col1 = start_f1 + n1
        ws.cell(row=2, column=avg_col1).value = "Avg"
        ws.cell(row=2, column=avg_col1).font = Font(bold=True)
        ws.cell(row=2, column=avg_col1).alignment = Alignment(horizontal='center')
        # Folder2 cycles
        cur2 = start_f2
        for i in range(n2):
            ws.cell(row=2, column=cur2 + i).value = f"Cycle{i+1}"
            ws.cell(row=2, column=cur2 + i).font = Font(bold=True)
            ws.cell(row=2, column=cur2 + i).alignment = Alignment(horizontal='center')  # center align cycle headers
        avg_col2 = start_f2 + n2
        ws.cell(row=2, column=avg_col2).value = "Avg"
        ws.cell(row=2, column=avg_col2).font = Font(bold=True)
        ws.cell(row=2, column=avg_col2).alignment = Alignment(horizontal='center')  # center align avg header

        # Diff header at row2
        ws.cell(row=2, column=diff_col).value = "Diff"
        ws.cell(row=2, column=diff_col).font = Font(bold=True)
        ws.cell(row=2, column=diff_col).alignment = Alignment(horizontal='center')  # center align diff header

        # write keys rows starting at row 3
        keys = []
        if d1:
            keys.extend(d1["keys"])
        if d2:
            for k in d2["keys"]:
                if k not in keys:
                    keys.append(k)

        row_idx = 3
        for key in keys:
            ws.cell(row=row_idx, column=1).value = key
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')  # wrap text in key column
            # folder1 values
            vals1 = d1["values"].get(key, []) if d1 else []
            # write cycles
            for i in range(n1):
                v = vals1[i] if i < len(vals1) else None
                ws.cell(row=row_idx, column=start_f1 + i).value = _format_number_for_cell(v)
                ws.cell(row=row_idx, column=start_f1 + i).alignment = Alignment(horizontal='center')  # align numbers
            # avg1
            present_vals1 = [v for v in vals1 if v is not None]
            avg1 = None
            if present_vals1:
                avg1 = sum(present_vals1) / len(present_vals1)
            ws.cell(row=row_idx, column=avg_col1).value = _format_number_for_cell(avg1)
            ws.cell(row=row_idx, column=avg_col1).alignment = Alignment(horizontal='center')  # center align numbers

            # folder2 values
            vals2 = d2["values"].get(key, []) if d2 else []
            for i in range(n2):
                v = vals2[i] if i < len(vals2) else None
                ws.cell(row=row_idx, column=start_f2 + i).value = _format_number_for_cell(v)
                ws.cell(row=row_idx, column=start_f2 + i).alignment = Alignment(horizontal='center')  # align numbers
            present_vals2 = [v for v in vals2 if v is not None]
            avg2 = None
            if present_vals2:
                avg2 = sum(present_vals2) / len(present_vals2)
            ws.cell(row=row_idx, column=avg_col2).value = _format_number_for_cell(avg2)
            ws.cell(row=row_idx, column=avg_col2).alignment = Alignment(horizontal='center')
            
            # diff
            diff = None
            if avg1 is not None and avg2 is not None:
                diff = avg1 - avg2
            ws.cell(row=row_idx, column=diff_col).value = _format_number_for_cell(diff)
            ws.cell(row=row_idx, column=diff_col).alignment = Alignment(horizontal='center')  # center align numbers
            if diff is not None and diff < 0.0 and key[0].isupper():
                ws.cell(row=row_idx, column=diff_col).fill = red_fill
            row_idx += 1

        # Auto-adjust column widths based on content
        try:
            # Auto-adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Check header rows (rows 1-2)
                for row in range(1, 3):
                    cell = ws.cell(row=row, column=column[0].column)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Check data rows (starting from row 3)
                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=column[0].column)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Set column width with some padding
                ws.column_dimensions[column_letter].width = max_length + 2
        except Exception:
            pass

    def create_excel_file(folder1, folder2, prefix1, prefix2, get_first_value=True, selected_keys = None):
        data1 = collect_folder_data(folder1, r"_start_", get_first_value)
        data2 = collect_folder_data(folder2, r"_start_", get_first_value)

        all_apps = sorted(set(list(data1.keys()) + list(data2.keys())))

        # Determine output filename
        if get_first_value:
            filename = f"CompareMemory_{prefix1}_{prefix2}_first.xlsx"
        else:
            filename = f"CompareMemory_{prefix1}_{prefix2}_re-entry.xlsx"
        
        out_path = os.path.join(folder1, filename)
        wb = Workbook()
        # remove default sheet
        default = wb.active
        
        # Check if we have any apps to process
        if not all_apps:
            # No apps found, create a placeholder sheet with message
            ws = wb.create_sheet(title="No Data")
            ws.cell(1, 1).value = "No memory data files found matching pattern '_start_'"
            ws.cell(1, 1).font = Font(bold=True)
            ws.cell(2, 1).value = f"Folder 1: {folder1}"
            ws.cell(3, 1).value = f"Folder 2: {folder2}"
            wb.remove(default)
            wb.save(out_path)
            print(f"[WARNING] No apps found in folders. Created placeholder Excel: {out_path}")
            return out_path
        
        wb.remove(default)
        # Create sheets for each app
        for app in all_apps:
            ws = wb.create_sheet(title=(app.capitalize()[:31]))  # sheet name max 31 chars
            d1 = data1.get(app)
            d2 = data2.get(app)
            write_app_sheet(ws, app, d1, d2, prefix1, prefix2)

        # Create summary sheet
        if selected_keys:
            missing_keys = [k for k in selected_keys if not any(
                (d1 and k in d1["keys"]) or (d2 and k in d2["keys"])
                for d1, d2 in [(data1.get(app), data2.get(app)) for app in all_apps]
            )]
            if missing_keys:
                print(f"[Error] Missing keys in data: {missing_keys}. Skip summary sheet")
            else:
                create_summary_sheet(wb, data1, data2, prefix1, prefix2, selected_keys)
                wb.move_sheet(wb["Summary"], offset=-len(wb.sheetnames)+1)
        
        wb.save(out_path)
        return out_path
    
    def create_summary_sheet(wb, data1, data2, prefix1, prefix2, selected_keys):
        ws = wb.create_sheet("Summary")
        red_fill = PatternFill(start_color="e80707", end_color="e80707", fill_type="solid")
        # Collect all apps
        
        all_apps = sorted(set(list(data1.keys()) + list(data2.keys())))
        # New.......
        #seen = set()
        # all_apps = []
        # for d in [data1, data2]:
        #     for app in d.keys():
        #         if app not in seen:
        #             all_apps.append(app)
        #             seen.add(app)

        # ---- Row1: merged headers for each key ----
        col = 2  # Column 1 reserved for App names
        for key in selected_keys:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
            ws.cell(row=1, column=col).value = key
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
            col += 3

        # ---- Row2: prefix1, prefix2, Diff ----
        col = 2
        for key in selected_keys:
            ws.cell(row=2, column=col).value = prefix1
            ws.cell(row=2, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")

            ws.cell(row=2, column=col + 1).value = prefix2
            ws.cell(row=2, column=col + 1).font = Font(bold=True)
            ws.cell(row=2, column=col + 1).alignment = Alignment(horizontal="center")

            ws.cell(row=2, column=col + 2).value = "Diff"
            ws.cell(row=2, column=col + 2).font = Font(bold=True)
            ws.cell(row=2, column=col + 2).alignment = Alignment(horizontal="center")

            col += 3

        # ---- Row3+: app names and averages ----
        row = 3
        for app in all_apps:
            ws.cell(row=row, column=1).value = app.capitalize()
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            col = 2
            for key in selected_keys:
                # Get averages for prefix1
                avg1 = None
                d1 = data1.get(app)
                if d1 and key in d1["values"]:
                    vals = [v for v in d1["values"][key] if v is not None]
                    if vals:
                        avg1 = sum(vals) / len(vals)

                # Get averages for prefix2
                avg2 = None
                d2 = data2.get(app)
                if d2 and key in d2["values"]:
                    vals = [v for v in d2["values"][key] if v is not None]
                    if vals:
                        avg2 = sum(vals) / len(vals)

                # Fill values
                ws.cell(row=row, column=col).value = _format_number_for_cell(avg1)
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=col + 1).value = _format_number_for_cell(avg2)
                ws.cell(row=row, column=col + 1).alignment = Alignment(horizontal="center")

                diff = None
                if avg1 is not None and avg2 is not None:
                    diff = avg1 - avg2
                ws.cell(row=row, column=col + 2).value = _format_number_for_cell(diff)
                if diff is not None and diff < 0.0:
                    ws.cell(row=row, column=col + 2).fill = red_fill
                ws.cell(row=row, column=col + 2).alignment = Alignment(horizontal="center")
                col += 3

            row += 1

        # Auto-adjust column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Check all rows in the column
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set column width with some padding
            ws.column_dimensions[column_letter].width = max_length + 2


    def create_start_end_excel_file(folder1, prefix1, get_first_value=True):
        #data_start = collect_folder_data(folder1, "_Start_", get_first_value)
        data_start = collect_folder_data(folder1, r"_start_", get_first_value)

        data_end = collect_folder_data(folder1, r"_end_", get_first_value)
        #data_end = collect_folder_data(folder1, "_End_", get_first_value)
        

        all_apps = sorted(set(list(data_start.keys()) + list(data_end.keys())))

        # Determine output filename
        if get_first_value:
            filename = f"CompareMemory_{prefix1}_StartEnd_first.xlsx"
        else:
            filename = f"CompareMemory_{prefix1}_StartEnd_re-entry.xlsx"
        
        out_path = os.path.join(folder1, filename)
        wb = Workbook()
        # remove default sheet
        default = wb.active
        
        # Check if we have any apps to process
        if not all_apps:
            # No apps found, create a placeholder sheet with message
            ws = wb.create_sheet(title="No Data")
            ws.cell(1, 1).value = "No memory data files found matching patterns '_start_' or '_end_'"
            ws.cell(1, 1).font = Font(bold=True)
            ws.cell(2, 1).value = f"Folder: {folder1}"
            wb.remove(default)
            wb.save(out_path)
            print(f"[WARNING] No apps found for Start/End comparison. Created placeholder Excel: {out_path}")
            return out_path

        wb.remove(default)

        # Create prefixes for Start and End
        prefix_start = f"{prefix1}-Start"
        prefix_end = f"{prefix1}-End"

        for app in all_apps:
            ws = wb.create_sheet(title=(app[:31]))  # sheet name max 31 chars
            d_start = data_start.get(app)
            d_end = data_end.get(app)
            write_app_sheet(ws, app, d_start, d_end, prefix_start, prefix_end)

        wb.save(out_path)
        return out_path


    # Main logic starts here
    # Normalize UNC paths and handle network access
    folder1 = os.path.normpath(dut_fog_folder_path)
    folder2 = os.path.normpath(ref_fog_folder_path)

    # Check if paths are accessible
    def is_accessible(path):
        try:
            return os.path.isdir(path)
        except Exception as e:
            raise Exception(f"Cannot access {path}: {str(e)}")

    if not is_accessible(folder1) or not is_accessible(folder2):
        raise Exception("Provided paths must be accessible folders")

    # Run
    prefix1 = get_prefix_from_folder(folder1)
    prefix2 = get_prefix_from_folder(folder2)

    # Create first Excel file (First entry)
    out_path1 = create_excel_file(folder1, folder2, prefix1, prefix2, get_first_value=True)
    
    # Create second Excel file (Re-en ryentry)
    out_path2 = create_excel_file(folder1, folder2, prefix1, prefix2, get_first_value=False)

    # Create third Excel file (Compare Start vs End first entry within same folder)
    out_path3 = create_start_end_excel_file(folder1, prefix1, get_first_value=True)

    # Create fourth Excel file (Compare Start vs End re-entry within same folder)
    out_path4 = create_start_end_excel_file(folder1, prefix1, get_first_value=False)

    # Select key
    selected_keys = ["MemFree", "MemAvailable"]

    # Create summary sheet 
    create_excel_file(folder1, folder2, prefix1, prefix2, get_first_value=True, selected_keys=selected_keys)
    create_excel_file(folder1, folder2, prefix1, prefix2, get_first_value=False, selected_keys=selected_keys)

    return out_path1, out_path2, out_path3, out_path4

def main():
    # Check number of arguments
    if len(sys.argv) != 3:
        print("Usage: python temp.py <ref_fog_folder_path> <dut_fog_folder_path>")
        sys.exit(1)
    
    # Get path
    dut_path = sys.argv[1]
    ref_path = sys.argv[2]
    
    # Check valid path
    if not os.path.isdir(ref_path):
        print(f"Error: Reference folder path does not exist: {ref_path}")
        sys.exit(1)
    
    if not os.path.isdir(dut_path):
        print(f"Error: DUT folder path does not exist: {dut_path}")
        sys.exit(1)
    
    # Main
    
    diff_memory(dut_path, ref_path)

if __name__ == "__main__":
    main()
