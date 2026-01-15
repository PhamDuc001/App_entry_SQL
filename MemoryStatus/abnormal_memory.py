"""
Optimized version of device_performance_analyzer.py
Performance improvements and code quality enhancements with comprehensive device analysis
"""

import os
import re
import sys
import zipfile
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Any
from dataclasses import dataclass
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

# Import the new app start/kill analyzer
try:
    # Try relative imports first (when used as module)
    from .app_start_kill_analyzer import AppStartKillAnalyzer, AppStartKillInfo, APP_PACKAGE_MAPPING, FOLDER_APP_PART_MAPPING
    from .analyze_pss import extract_pss_for_package
except ImportError:
    # Fall back to absolute imports (when run directly)
    from app_start_kill_analyzer import AppStartKillAnalyzer, AppStartKillInfo, APP_PACKAGE_MAPPING, FOLDER_APP_PART_MAPPING
    from analyze_pss import extract_pss_for_package


# Configuration constants
class Config:
    UPTIME_THRESHOLD_MINUTES = 10
    STATUS_OK = "OK"
    STATUS_NG = "NG"
    COLUMN_WIDTHS = {
        1: 40,  # File name
        2: 20,  # Uptime
        3: 15,  # Status
        4: 80   # Raw log
    }
    COLORS = {
        'header': "366092",
        'ok': "C6EFCE",
        'ng': "FFC7CE"
    }
    MAX_RAW_LOG_LENGTH = 100
    FILE_BUFFER_SIZE = 65536
    PSS_DIFF_THRESHOLD = 10  # MB threshold for PSS difference reporting
    IO_DIFF_THRESHOLD = 600   # MB threshold for I/O difference highlighting
    DEFAULT_RAM_SIZE_GB = 8   # Default RAM size in GB
    DEFAULT_PSS_THRESHOLD_MB = 800  # Default PSS threshold in MB for 8GB RAM
    RAM_TO_PSS_THRESHOLD_MAP = {
        6: 100,   # 500MB threshold for devices with less than 6GB RAM
        8: 100,   # 800MB threshold for devices with 6-8GB RAM
        12: 100, # 1GB threshold for devices with more than 8GB RAM (using 12GB as key for 8GB+ devices)
    }
    
    @classmethod
    def get_threshold_for_ram(cls, ram_size_gb):
        """
        Determine PSS threshold based on device RAM size using the mapping in Config.
        
        Args:
            ram_size_gb (int): Device RAM size in GB
            
        Returns:
            int: Threshold in MB
        """
        # Sort the RAM sizes in descending order to check from highest to lowest
        sorted_ram_sizes = sorted(cls.RAM_TO_PSS_THRESHOLD_MAP.keys(), reverse=True)
        
        # Find the appropriate threshold based on RAM size
        for ram_size in sorted_ram_sizes:
            if ram_size_gb >= ram_size:
                return cls.RAM_TO_PSS_THRESHOLD_MAP[ram_size]
        
        # If RAM size is smaller than the smallest key in the map, return the threshold for the smallest key
        smallest_ram_size = min(cls.RAM_TO_PSS_THRESHOLD_MAP.keys())
        return cls.RAM_TO_PSS_THRESHOLD_MAP[smallest_ram_size]


# Pre-compiled regex patterns for performance
UPTIME_PATTERN = re.compile(r"Uptime:\s+up\s+(.+?),\s+load average:")

# FATAL/ANR patterns
ANR_PATTERN = re.compile(r"ANR in (.+?)(?:\s+\(pid\s+(\d+)\))?")
FATAL_PATTERN = re.compile(r"FATAL EXCEPTION: (.+?)(?:\s+pid\s+(\d+))?")

# I/O patterns
IO_PATTERN = re.compile(r"(Read_top|Write_top)\(KB\):\s*(.*)")
IO_PROCESS_PATTERN = re.compile(r"([^(]+)\((?:pid\s+\d+,)?\s*[\d,]+K\)")
IO_VALUE_PATTERN = re.compile(r"\d+")


@dataclass
class UptimeData:
    """Data class for uptime information"""
    filename: str
    uptime_minutes: Optional[int]
    status: str
    raw_line: str
    extracted_file_path: Optional[Path] = None
    io_read_data: Optional[List[Tuple[str, float]]] = None
    io_write_data: Optional[List[Tuple[str, float]]] = None
    part_name: Optional[str] = None  # Added to track which part this file belongs to


@dataclass
class CrashData:
    """Data class for FATAL/ANR crash information"""
    filename: str
    crash_type: str  # "ANR" or "FATAL"
    app_name: str
    pid: Optional[str]
    raw_line: str


@dataclass
class AnalysisResult:
    """Data class for analysis results"""
    prefix: str
    uptime_data: List[UptimeData]
    crash_data: List[CrashData]
    ok_count: int
    ng_count: int
    total_count: int
    anr_count: int
    fatal_count: int
    # Add app start/kill information
    app_start_kill_data: List[AppStartKillInfo] = None
    # Add field to store averaged IO data
    averaged_io_data: Optional[Dict[str, Dict[str, List[Tuple[str, float]]]]] = None
    # Add field to store dumpstate file contents
    dumpstate_contents: Optional[Dict[Path, str]] = None
    # Add field to store compiler information
    compiler_data: Optional[Dict[str, str]] = None
    
    def __post_init__(self):
        if self.app_start_kill_data is None:
            self.app_start_kill_data = []
        if self.averaged_io_data is None:
            self.averaged_io_data = {}
        if self.dumpstate_contents is None:
            self.dumpstate_contents = {}
        if self.compiler_data is None:
            self.compiler_data = {}


@dataclass
class ComparisonResult:
    """Data class for device comparison results"""
    dut_result: AnalysisResult
    ref_result: AnalysisResult
    uptime_comparison: Dict[str, Any]
    crash_comparison: Dict[str, Any]
    io_comparison: Dict[str, Any]
    app_comparison: Dict[str, Any]


class Device:
    """Base class representing a device to be analyzed"""
    
    def __init__(self, name: str, folder_path: Path, config: Config):
        self.name = name  # "DUT" or "REF"
        self.folder_path = Path(folder_path)
        self.config = config
        self._ram_size = 0  # RAM size in GB
        self._debug_level = "UNKNOWN"  # Debug level as string
        self._anr_fatal = "UNKNOWN"  # ANR/FATAL status as string
        self._uptime = "UNKNOWN"  # Uptime status as string
        self.analysis_result: Optional[AnalysisResult] = None
        self.app_analyzer = AppStartKillAnalyzer() if name == "DUT" else AppStartKillAnalyzer()
    
    # Getter and setter methods for ram_size
    def get_ram_size(self) -> int:
        """Get the RAM size of the device in GB."""
        return self._ram_size
    
    def set_ram_size(self, ram_size: int) -> None:
        """Set the RAM size of the device in GB."""
        if not isinstance(ram_size, int) or ram_size < 0:
            raise ValueError("RAM size must be a non-negative integer.")
        self._ram_size = ram_size
    
    # Getter and setter methods for debug_level
    def get_debug_level(self) -> str:
        """Get the debug level of the device."""
        return self._debug_level
    
    def set_debug_level(self, debug_level: str) -> None:
        """Set the debug level of the device."""
        if not isinstance(debug_level, str):
            raise ValueError("Debug level must be a string.")
        self._debug_level = debug_level
    
    # Getter and setter methods for anr_fatal
    def get_anr_fatal(self) -> str:
        """Get the ANR/FATAL status of the device."""
        return self._anr_fatal
    
    def set_anr_fatal(self, anr_fatal: str) -> None:
        """Set the ANR/FATAL status of the device."""
        if not isinstance(anr_fatal, str):
            raise ValueError("ANR/FATAL status must be a string.")
        self._anr_fatal = anr_fatal
    
    # Getter and setter methods for uptime
    def get_uptime(self) -> str:
        """Get the uptime status of the device."""
        return self._uptime
    
    def set_uptime(self, uptime: str) -> None:
        """Set the uptime status of the device."""
        if not isinstance(uptime, str):
            raise ValueError("Uptime status must be a string.")
        self._uptime = uptime
    
    def analyze(self, extracted: bool = False) -> AnalysisResult:
        """Perform comprehensive analysis of this device"""
        # Find the largest dumpstate file to determine RAM size and debug level
        dumpstate_file = None
        largest_size = 0
        
        if self.folder_path.exists():
            for dirpath, dirnames, filenames in os.walk(self.folder_path):
                for filename in filenames:
                    if filename.startswith('dumpstate-') and filename.endswith('.txt'):
                        file_path = os.path.join(dirpath, filename)
                        try:
                            file_size = os.path.getsize(file_path)
                            if file_size > largest_size:
                                largest_size = file_size
                                dumpstate_file = file_path
                        except Exception as e:
                            print(f"Error getting size for file {file_path}: {e}")
        
        # Set RAM size and debug level for this device
        if dumpstate_file:
            try:
                with open(dumpstate_file, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                # Import the functions here to avoid circular imports
                try: 
                    from .analyze_pss import get_ram_size, get_debug_level
                    # Read the entire file to detect RAM size and debug level
                except ImportError:
                    from analyze_pss import get_ram_size, get_debug_level
                ram_size = get_ram_size(content)
                debug_level = get_debug_level(content)
                self.set_ram_size(ram_size)
                self.set_debug_level(debug_level)
                print(f"{self.name} RAM size: {ram_size}GB, Debug level: {debug_level}")
            except Exception as e:
                print(f"Error reading dumpstate file for {self.name} RAM size and debug level detection: {e}")
        
        analyzer = DevicePerformanceAnalyzer(self.config)
        self.analysis_result = analyzer.analyze_folder(self.folder_path, extracted)
        
        return self.analysis_result
    
    def get_uptime_data(self) -> List[UptimeData]:
        """Get uptime data for this device"""
        if self.analysis_result:
            return self.analysis_result.uptime_data
        return []
    
    def get_crash_data(self) -> List[CrashData]:
        """Get crash data for this device"""
        if self.analysis_result:
            return self.analysis_result.crash_data
        return []
    
    def get_prefix(self) -> str:
        """Get prefix for this device"""
        if self.analysis_result:
            return self.analysis_result.prefix
        # Fallback to analyzer method
        analyzer = DevicePerformanceAnalyzer(self.config)
        return analyzer.get_prefix(self.folder_path)


class DUT(Device):
    """Device Under Test - specific implementation if needed"""
    
    def __init__(self, folder_path: Path, config: Config):
        super().__init__("DUT", folder_path, config)


class REF(Device):
    """Reference Device - specific implementation if needed"""
    
    def __init__(self, folder_path: Path, config: Config):
        super().__init__("REF", folder_path, config)


class DeviceComparator:
    """Handles comparison between two devices"""
    
    def __init__(self, dut: Device, ref: Device, config: Config):
        self.dut = dut
        self.ref = ref
        self.config = config
        self.analyzer = DevicePerformanceAnalyzer(config)
    
    def compare(self) -> ComparisonResult:
        """Compare the two devices and return results"""
        # Perform analysis on both devices if not already done
        if not self.dut.analysis_result:
            self.dut.analyze(extracted=True)
            
        if not self.ref.analysis_result:
            self.ref.analyze(extracted=True)
        
        # Set anr_fatal and uptime based on comparison results
        # If no NG -> set anr_fatal to OK otherwise set Abnormal
        if self.dut.analysis_result.ng_count == 0:
            self.dut.set_anr_fatal("OK")
        else:
            self.dut.set_anr_fatal("Abnormal")
            
        if self.ref.analysis_result.ng_count == 0:
            self.ref.set_anr_fatal("OK")
        else:
            self.ref.set_anr_fatal("Abnormal")
            
        # Same for uptime
        if self.dut.analysis_result.ng_count == 0:
            self.dut.set_uptime("OK")
        else:
            self.dut.set_uptime("Abnormal")
            
        if self.ref.analysis_result.ng_count == 0:
            self.ref.set_uptime("OK")
        else:
            self.ref.set_uptime("Abnormal")
        
        # Create comparison result
        comparison_result = ComparisonResult(
            dut_result=self.dut.analysis_result,
            ref_result=self.ref.analysis_result,
            uptime_comparison=self._compare_uptime(),
            crash_comparison=self._compare_crashes(),
            io_comparison=self._compare_io(),
            app_comparison=self._compare_apps()
        )
        
        return comparison_result
    
    def _compare_uptime(self) -> Dict[str, Any]:
        """Compare uptime data between devices"""
        # Placeholder for actual comparison logic
        return {"status": "compared"}
    
    def _compare_crashes(self) -> Dict[str, Any]:
        """Compare crash data between devices"""
        # Placeholder for actual comparison logic
        return {"status": "compared"}
    
    def _compare_io(self) -> Dict[str, Any]:
        """Compare I/O data between devices"""
        # Placeholder for actual comparison logic
        return {"status": "compared"}
    
    def _compare_apps(self) -> Dict[str, Any]:
        """Compare app start/kill data between devices"""
        # Placeholder for actual comparison logic
        return {"status": "compared"}
    
    def generate_excel_report(self, output_path: Path) -> bool:
        """Generate Excel comparison report"""
        if not self.dut.analysis_result:
            self.dut.analyze(extracted=True)
            
        if not self.ref.analysis_result:
            self.ref.analyze(extracted=True)
            
        return self.analyzer.generate_excel_report(
            self.dut.analysis_result, 
            self.ref.analysis_result, 
            output_path
        )
    
    def generate_console_report(self) -> str:
        """Generate console summary report"""
        if not self.dut.analysis_result:
            self.dut.analyze(extracted=True)
            
        if not self.ref.analysis_result:
            self.ref.analyze(extracted=True)
            
        ret = ""
        ret_console = "\n=== UPTIME STATUS SUMMARY ==="
        ret_console += f"\n{self.dut.analysis_result.prefix}: OK = {self.dut.analysis_result.ok_count}, NG = {self.dut.analysis_result.ng_count}"
        ret_console += f"\n{self.ref.analysis_result.prefix}: OK = {self.ref.analysis_result.ok_count}, NG = {self.ref.analysis_result.ng_count}"
        
        if self.dut.analysis_result.ng_count == 0 and self.ref.analysis_result.ng_count == 0:
            ret_console += f"\n✅ All uptime < {self.config.UPTIME_THRESHOLD_MINUTES} minutes"
        else:
            NG_result = (f"\n⚠️  PRE-SETTING HAVE ISSUES - "
                   f"Some uptime > {self.config.UPTIME_THRESHOLD_MINUTES} minutes")
            ret = "\n=== UPTIME STATUS SUMMARY ==="
            ret += NG_result
            ret_console += NG_result
        
        # Add crash summary
        ret_console += "\n\n=== CRASH ANALYSIS SUMMARY ==="
        ret_console += f"\n{self.dut.analysis_result.prefix}: ANR = {self.dut.analysis_result.anr_count}, FATAL = {self.dut.analysis_result.fatal_count}"
        ret_console += f"\n{self.ref.analysis_result.prefix}: ANR = {self.ref.analysis_result.anr_count}, FATAL = {self.ref.analysis_result.fatal_count}"
        
        total_crashes1 = self.dut.analysis_result.anr_count + self.dut.analysis_result.fatal_count
        total_crashes2 = self.ref.analysis_result.anr_count + self.ref.analysis_result.fatal_count
        
        if total_crashes1 == 0 and total_crashes2 == 0:
            ret_console += "\n✅ No crashes detected in either device"
        else:
            NG_result = f"\n⚠️  CRASHES DETECTED - Total: {total_crashes1} (DUT), {total_crashes2} (REF)"
            ret_console += NG_result
            ret += "\n\n=== CRASH ANALYSIS SUMMARY ==="
            ret += NG_result
        
        return ret


class DevicePerformanceAnalyzer:
    """Optimized device performance analyzer with improved performance and error handling"""
    
    def __init__(self, config: Config = None):
        self.config = config or Config()
    
    def extract_largest_file_from_zip(self, zip_path: Path, extract_dir: Path) -> Optional[Path]:
        """Extract largest file from zip with intelligent caching"""
        extract_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate cache filename
        zip_mtime = zip_path.stat().st_mtime
        cache_filename = f"{zip_path.stem}_{int(zip_mtime)}.txt"
        cache_path = extract_dir / cache_filename
        
        # Return cached file if exists
        if cache_path.exists():
            return cache_path
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                infos = z.infolist()
                if not infos:
                    return None
                
                largest = max(infos, key=lambda x: x.file_size)
                
                # Extract to cache location
                with open(cache_path, "wb") as f:
                    f.write(z.read(largest))
                
                return cache_path
                
        except (zipfile.BadZipFile, OSError) as e:
            print(f"Error extracting {zip_path}: {e}")
            return None
    
    def parse_file_content(self, file_path: Path) -> Tuple[
        Tuple[Optional[int], str, str],  # uptime data
        List[Tuple[str, float]],         # io read data
        List[Tuple[str, float]],         # io write data
        List[CrashData]                  # crash data
    ]:
        """Parse all required information from log file in a single pass"""
        # Initialize return values
        uptime_data = (None, self.config.STATUS_NG, "No uptime line found")
        process_reads = defaultdict(int)  # For I/O read data
        process_writes = defaultdict(int)  # For I/O write data
        crashes = []
        
        try:

            with open(file_path, "r", encoding="utf-8", 
                     errors="ignore", buffering=self.config.FILE_BUFFER_SIZE) as f:
                for line in f:
                    # Check for uptime
                    if uptime_data[0] is None:  # Only search until we find uptime
                        uptime_match = UPTIME_PATTERN.search(line)
                        if uptime_match:
                            uptime_str = uptime_match.group(1)
                            uptime_minutes = self._convert_uptime_to_minutes(uptime_str)
                            status = (self.config.STATUS_NG 
                                    if uptime_minutes > self.config.UPTIME_THRESHOLD_MINUTES 
                                    else self.config.STATUS_OK)
                            uptime_data = (uptime_minutes, status, line.strip())
                    
                    # Check for I/O lines (Read_top or Write_top)
                    io_match = IO_PATTERN.search(line)
                    if io_match:
                        io_type = io_match.group(1)
                        processes_info = io_match.group(2).strip()
                        # Split by spaces to get process name and value pairs
                        parts = processes_info.split()
                        
                        # Process the pairs more efficiently
                        i = 0
                        target_dict = process_reads if io_type == "Read_top" else process_writes
                        while i < len(parts) - 1:  # -1 because we need to check i+1
                            # Check if this is a process entry
                            if '(' in parts[i] and ')' in parts[i]:
                                try:
                                    # Get process name (before the PID in parentheses)
                                    process_name = parts[i].split('(')[0]
                                    # Get value (next part)
                                    value = int(parts[i + 1])
                                    # Add to the appropriate process total
                                    target_dict[process_name] += value
                                    i += 2  # Move to next pair
                                except ValueError:
                                    # Skip if we can't parse the value
                                    i += 1
                            else:
                                i += 1
                    
                    # Check for ANR
                    anr_match = ANR_PATTERN.search(line)
                    if anr_match:
                        app_name = anr_match.group(1).strip()
                        pid = anr_match.group(2) if len(anr_match.groups()) > 1 else None
                        crashes.append(CrashData(
                            filename=file_path.name,
                            crash_type="ANR",
                            app_name=app_name,
                            pid=pid,
                            raw_line=line.strip()
                        ))
                    
                    # Check for FATAL EXCEPTION
                    fatal_match = FATAL_PATTERN.search(line)
                    if fatal_match:
                        app_name = fatal_match.group(1).strip()
                        pid = fatal_match.group(2) if len(fatal_match.groups()) > 1 else None
                        crashes.append(CrashData(
                            filename=file_path.name,
                            crash_type="FATAL",
                            app_name=app_name,
                            pid=pid,
                            raw_line=line.strip()
                        ))
        
        except (IOError, OSError) as e:
            print(f"Error reading file {file_path}: {e}")
            return (
                (None, self.config.STATUS_NG, f"Error reading file: {e}"),
                [],
                [],
                crashes
            )
        
        # Convert I/O read data to MB and get top 10
        # Since we only show top 10 in the report, we don't need to sort all entries
        # Use heapq.nlargest for better performance
        import heapq
        process_reads_mb = []
        for process, read_kb in process_reads.items():
            read_mb = read_kb / 1024.0
            process_reads_mb.append((read_mb, process))
        
        # Get top 10 processes by read MB
        top_10_reads = heapq.nlargest(10, process_reads_mb, key=lambda x: x[0])
        # Convert read_mb to integer to avoid decimal places in the report
        process_reads_mb = [(process, int(read_mb)) for read_mb, process in top_10_reads]
        
        # Convert I/O write data to MB and get top 10
        # Since we only show top 10 in the report, we don't need to sort all entries
        # Use heapq.nlargest for better performance
        process_writes_mb = []
        for process, write_kb in process_writes.items():
            write_mb = write_kb / 1024.0
            process_writes_mb.append((write_mb, process))
        
        # Get top 10 processes by write MB
        top_10_writes = heapq.nlargest(10, process_writes_mb, key=lambda x: x[0])
        # Convert write_mb to integer to avoid decimal places in the report
        process_writes_mb = [(process, int(write_mb)) for write_mb, process in top_10_writes]
        
        return uptime_data, process_reads_mb, process_writes_mb, crashes
    
    def _convert_uptime_to_minutes(self, uptime_str: str) -> int:
        """Convert uptime string to total minutes with pre-compiled patterns"""
        total_minutes = 0
        
        # Parse all time components at once using a single regex
        time_pattern = re.compile(r'(\d+)\s+(weeks?|days?|hours?|minutes?)')
        for match in time_pattern.finditer(uptime_str):
            value = int(match.group(1))
            unit = match.group(2)
            
            if unit.startswith('week'):
                total_minutes += value * 7 * 24 * 60
            elif unit.startswith('day'):
                total_minutes += value * 24 * 60
            elif unit.startswith('hour'):
                total_minutes += value * 60
            elif unit.startswith('minute'):
                total_minutes += value
        
        return total_minutes
    
    def extract_all_zips(self, folder: Path) -> Dict[Path, Path]:
        """Extract all zip files first and return mapping of zip files to extracted file paths"""
        cache_dir = folder / "_tmp"
        cache_dir.mkdir(parents=True, exist_ok=True)
        
        zip_files = [f for f in folder.glob("*.zip") if f.is_file()]
        zip_to_extracted = {}
        
        # Use threading to extract files in parallel
        with ThreadPoolExecutor(max_workers=8) as executor:
            # Submit all extraction tasks
            future_to_zip = {
                executor.submit(self.extract_largest_file_from_zip, zip_file, cache_dir): zip_file 
                for zip_file in sorted(zip_files)
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_zip):
                zip_file = future_to_zip[future]
                try:
                    dump_path = future.result()
                    if dump_path:
                        zip_to_extracted[zip_file] = dump_path
                except Exception as e:
                    print(f"Error extracting {zip_file}: {e}")
        
        return zip_to_extracted
    
    def collect_all_data_from_zips(self, folder: Path) -> Tuple[List[UptimeData], List[CrashData]]:
        """Collect all data (uptime and crashes) from zip files in a single pass per file"""
        uptime_data = []
        crash_data = []
        
        # Extract all zip files first
        zip_to_extracted = self.extract_all_zips(folder)
        
        # Then process all extracted files
        for zip_file, dump_path in zip_to_extracted.items():
            # Parse all content in a single pass directly
            uptime_result, io_read_data, io_write_data, file_crash_data = self.parse_file_content(dump_path)
            uptime_minutes, status, raw_line = uptime_result
            
            # Extract part name from zip file name
            part_name = self._extract_part_name(zip_file.name)
            
            uptime_data.append(UptimeData(
                filename=zip_file.name,
                uptime_minutes=uptime_minutes,
                status=status,
                raw_line=raw_line,
                extracted_file_path=dump_path,
                io_read_data=io_read_data,
                io_write_data=io_write_data,
                part_name=part_name
            ))
            
            crash_data.extend(file_crash_data)
        
        return uptime_data, crash_data
    
    def collect_all_data_from_extracted(self, folder: Path) -> Tuple[List[UptimeData], List[CrashData]]:
        """Collect all data (uptime and crashes) from extracted folders in a single pass per file"""
        uptime_data = []
        crash_data = []
        
        # Group subdirectories by part name using os.walk to handle subfolders
        part_groups = defaultdict(list)
        for dirpath, dirnames, filenames in os.walk(folder):
            # Convert dirpath to Path object for consistency
            current_dir = Path(dirpath)
            
            # Only process directories that are direct subdirectories of the main folder
            # or have a valid part name in their path
            if current_dir != folder:
                part_name = self._extract_part_name(current_dir.name)
                if part_name:
                    part_groups[part_name].append(current_dir)
        # Process all folders for each part
        for part_name, sub_dirs in part_groups.items():
            for sub_dir in sub_dirs:
                largest_file = self._find_largest_file(sub_dir)
                if not largest_file:
                    continue
                
                # Parse all content in a single pass directly
                uptime_result, io_read_data, io_write_data, file_crash_data = self.parse_file_content(largest_file)
                uptime_minutes, status, raw_line = uptime_result
                
                uptime_data.append(UptimeData(
                    filename=sub_dir.name,
                    uptime_minutes=uptime_minutes,
                    status=status,
                    raw_line=raw_line,
                    extracted_file_path=largest_file,
                    io_read_data=io_read_data,
                    io_write_data=io_write_data,
                    part_name=part_name
                ))
                
                crash_data.extend(file_crash_data)
        
        return uptime_data, crash_data
    
    def _find_largest_file(self, directory: Path) -> Optional[Path]:
        """Find largest file in directory using os.walk for better performance"""
        try:
            largest_file = None
            largest_size = 0
            for dirpath, _, filenames in os.walk(directory):
                for filename in filenames:
                    file_path = Path(dirpath) / filename
                    try:
                        file_size = file_path.stat().st_size
                        if file_size > largest_size:
                            largest_size = file_size
                            largest_file = file_path
                    except (OSError, ValueError):
                        # Skip files that can't be accessed or have invalid sizes
                        continue
            return largest_file
        except (ValueError, OSError):
            return None
    
    def _extract_part_name(self, folder_name: str) -> Optional[str]:
        """Extract part name from folder name (e.g., 1part, 2part, Part1, Part2, etc.)"""
        # Look for patterns like 1part, 2part, Part1, Part2, etc. in the folder name
        # Handle both folder names and zip file names
        part_pattern = re.compile(r'((?:\d+part|part\d+))', re.IGNORECASE)
        match = part_pattern.search(folder_name)
        if match:
            # Normalize to lowercase format (e.g., "1part", "2part")
            part_name = match.group(1).lower()
            # Ensure consistent format: number before "part"
            if part_name.startswith('part'):
                # Convert "part1" to "1part"
                number = re.search(r'\d+', part_name)
                if number:
                    return f"{number.group()}part"
            return part_name

        # For zip files, look for the part pattern before _Bugreport.zip
        zip_pattern = re.compile(r'((?:\d+part|part\d+))_Bugreport\.zip$', re.IGNORECASE)
        match = zip_pattern.search(folder_name)
        if match:
            # Normalize to lowercase format (e.g., "1part", "2part")
            part_name = match.group(1).lower()
            # Ensure consistent format: number before "part"
            if part_name.startswith('part'):
                # Convert "part1" to "1part"
                number = re.search(r'\d+', part_name)
                if number:
                    return f"{number.group()}part"
            return part_name
        
        return None
    
    def _group_and_calculate_averages(self, uptime_data: List[UptimeData]) -> Dict[str, Dict[str, List[Tuple[str, float]]]]:
        """Group uptime data by parts and calculate averages for IO READ/WRITE using heapq.nlargest for better performance"""
        import heapq
        
        # Group data by part name
        part_data = defaultdict(list)
        for item in uptime_data:
            part_name = item.part_name if item.part_name else "unknown"
            # Skip "unknown" parts
            if part_name != "unknown":
                part_data[part_name].append(item)
        
        # Calculate averages for each part
        part_averages = {}
        for part_name, items in part_data.items():
            # Aggregate IO READ data
            process_read_totals = defaultdict(list)
            for item in items:
                if item.io_read_data:
                    for process, value in item.io_read_data:
                        process_read_totals[process].append(value)
            
            # Calculate average for each process and get top 10 using a min-heap
            avg_read_data = []
            for process, values in process_read_totals.items():
                avg_value = sum(values) / len(values)
                # Use a min-heap to keep the largest 10 values
                if len(avg_read_data) < 10:
                    heapq.heappush(avg_read_data, (avg_value, process))
                elif avg_value > avg_read_data[0][0]:
                    heapq.heapreplace(avg_read_data, (avg_value, process))
            
            # Convert heap to list and sort in descending order
            avg_read_data = [(process, avg_value) for avg_value, process in avg_read_data]
            avg_read_data.sort(key=lambda x: x[1], reverse=True)
            
            # Aggregate IO WRITE data
            process_write_totals = defaultdict(list)
            for item in items:
                if item.io_write_data:
                    for process, value in item.io_write_data:
                        process_write_totals[process].append(value)
            
            # Calculate average for each process and get top 10 using a min-heap
            avg_write_data = []
            for process, values in process_write_totals.items():
                avg_value = sum(values) / len(values)
                # Use a min-heap to keep the largest 10 values
                if len(avg_write_data) < 10:
                    heapq.heappush(avg_write_data, (avg_value, process))
                elif avg_value > avg_write_data[0][0]:
                    heapq.heapreplace(avg_write_data, (avg_value, process))
            
            # Convert heap to list and sort in descending order
            avg_write_data = [(process, avg_value) for avg_value, process in avg_write_data]
            avg_write_data.sort(key=lambda x: x[1], reverse=True)
            
            part_averages[part_name] = {
                'read': avg_read_data,
                'write': avg_write_data
            }
        
        return part_averages
    
    def _update_uptime_data_with_averages(self, uptime_data: List[UptimeData], part_io_data: Dict[str, Dict[str, Tuple[List[Tuple[str, float]], List[Tuple[str, float]]]]]) -> List[UptimeData]:
        """Update uptime data with averaged IO data"""
        # Create a mapping of part_name to representative item for each part
        part_representatives = {}
        for item in uptime_data:
            part_name = item.part_name if item.part_name else "unknown"
            if part_name not in part_representatives:
                # Create a new UptimeData item for this part
                part_representatives[part_name] = UptimeData(
                    filename=f"{part_name}_average",
                    uptime_minutes=item.uptime_minutes,
                    status=item.status,
                    raw_line=item.raw_line,
                    extracted_file_path=item.extracted_file_path,
                    io_read_data=part_io_data.get(part_name, {}).get('read', []),
                    io_write_data=part_io_data.get(part_name, {}).get('write', []),
                    part_name=part_name
                )
        
        # Return the representative items for each part
        return list(part_representatives.values())
    
    def get_prefix(self, folder: Path) -> str:
        """Extract prefix from folder/zip name using pathlib"""
        # Check for zip files first
        zip_files = list(folder.glob("*.zip"))
        if zip_files:
            tokens = re.split(r"[_\-]", zip_files[0].stem)
            if len(tokens) >= 3:
                return f"{tokens[0]}-{tokens[1]}-{tokens[2]}"
        
        # Check for subdirectories
        sub_dirs = [d for d in folder.iterdir() if d.is_dir()]
        if sub_dirs:
            tokens = re.split(r"[_\-]", sub_dirs[0].name)
            if len(tokens) >= 3:
                return f"{tokens[0]}-{tokens[1]}-{tokens[2]}"
        
        return "UNKNOWN"
    
    def analyze_folder(self, folder: Path, extracted: bool = False) -> AnalysisResult:
        """Analyze a single folder and return structured results"""
        prefix = self.get_prefix(folder)
        
        if extracted:
            uptime_data, crash_data = self.collect_all_data_from_extracted(folder)
        else:
            uptime_data, crash_data = self.collect_all_data_from_zips(folder)
        
        # Keep original uptime data for uptime sheets (individual files)
        original_uptime_data = uptime_data[:]
        
        # Group uptime data by parts and calculate averages for IO READ/WRITE
        part_io_data = self._group_and_calculate_averages(uptime_data)
        
        # Update uptime_data with averaged IO data for IO sheets and comparison sheets
        updated_uptime_data = self._update_uptime_data_with_averages(uptime_data, part_io_data)
        
        # Analyze app start/kill events
        app_start_kill_data = []
        if extracted:
            app_analyzer = AppStartKillAnalyzer()
            # Group subdirectories by part name for app analysis
            part_groups = defaultdict(list)
            for sub_dir in sorted([d for d in folder.iterdir() if d.is_dir()]):
                part_name = self._extract_part_name(sub_dir.name)
                if part_name:
                    part_groups[part_name].append(sub_dir)
            
            # Process all folders for each part
            for part_name, sub_dirs in part_groups.items():
                for sub_dir in sub_dirs:
                    app_info_list = app_analyzer.analyze_folder(sub_dir, part_name)
                    app_start_kill_data.extend(app_info_list)
        
        # Store dumpstate file contents for PSS analysis
        dumpstate_contents = {}
        for item in original_uptime_data:
            if item.extracted_file_path:
                try:
                    with open(item.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                        dumpstate_contents[item.extracted_file_path] = f.read()
                except Exception as e:
                    print(f"Error reading dumpstate file {item.extracted_file_path}: {e}")
        
        # Extract compiler information
        compiler_data = self._extract_compiler_info(dumpstate_contents)
        
        ok_count = sum(1 for item in original_uptime_data if item.status == self.config.STATUS_OK)
        ng_count = sum(1 for item in original_uptime_data if item.status == self.config.STATUS_NG)
        anr_count = sum(1 for crash in crash_data if crash.crash_type == "ANR")
        fatal_count = sum(1 for crash in crash_data if crash.crash_type == "FATAL")
        
        return AnalysisResult(
            prefix=prefix,
            uptime_data=original_uptime_data,  # Keep original uptime data for uptime sheets
            crash_data=crash_data,
            ok_count=ok_count,
            ng_count=ng_count,
            total_count=len(original_uptime_data),
            anr_count=anr_count,
            fatal_count=fatal_count,
            app_start_kill_data=app_start_kill_data,
            averaged_io_data=part_io_data,
            dumpstate_contents=dumpstate_contents,
            compiler_data=compiler_data
        )
    
    def generate_excel_report(self, result1: AnalysisResult, result2: AnalysisResult, 
                            output_path: Path) -> bool:
        """Generate optimized Excel report with crash analysis sheets"""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Create uptime sheets
            ws_dut = wb.create_sheet("DUT_Uptime")
            ws_ref = wb.create_sheet("REF_Uptime")
            
            # Create crash sheets
            ws_dut_crashes = wb.create_sheet("DUT_Crashes")
            ws_ref_crashes = wb.create_sheet("REF_Crashes")
            
            # Create comparison sheets for IO READ/WRITE
            ws_io_read_comparison = wb.create_sheet("IO_READ_COMPARISON")
            ws_io_write_comparison = wb.create_sheet("IO_WRITE_COMPARISON")
            
            # Create app start/kill analysis sheet
            ws_app_analysis = wb.create_sheet("App_Start_Kill_Analysis")
            
            # Create testing app PSS comparison sheet
            ws_testing_app_pss = wb.create_sheet("Testing_App_PSS_Comparison")
            
            # Create PSS analysis sheet (merged from separate PSS report)
            ws_pss_analysis = wb.create_sheet("PSS_Analysis")
            
            # Populate uptime sheets (using original uptime data)
            self._create_device_sheet(ws_dut, result1, "DUT")
            self._create_device_sheet(ws_ref, result2, "REF")
            
            # Populate crash sheets
            self._create_crash_sheet(ws_dut_crashes, result1, "DUT")
            self._create_crash_sheet(ws_ref_crashes, result2, "REF")
            
            # For comparison sheets, we need to use the averaged data
            # Use pre-calculated averaged data if available to avoid recalculation
            if result1.averaged_io_data and result2.averaged_io_data:
                # Create temporary AnalysisResult objects with pre-calculated averaged data
                averaged_result1 = AnalysisResult(
                    prefix=result1.prefix,
                    uptime_data=self._update_uptime_data_with_averages(result1.uptime_data, result1.averaged_io_data),
                    crash_data=result1.crash_data,
                    ok_count=result1.ok_count,
                    ng_count=result1.ng_count,
                    total_count=result1.total_count,
                    anr_count=result1.anr_count,
                    fatal_count=result1.fatal_count,
                    app_start_kill_data=result1.app_start_kill_data
                )
                averaged_result2 = AnalysisResult(
                    prefix=result2.prefix,
                    uptime_data=self._update_uptime_data_with_averages(result2.uptime_data, result2.averaged_io_data),
                    crash_data=result2.crash_data,
                    ok_count=result2.ok_count,
                    ng_count=result2.ng_count,
                    total_count=result2.total_count,
                    anr_count=result2.anr_count,
                    fatal_count=result2.fatal_count,
                    app_start_kill_data=result2.app_start_kill_data
                )
            else:
                # If averaged data is not available, calculate it
                averaged_result1 = AnalysisResult(
                    prefix=result1.prefix,
                    uptime_data=self._get_averaged_uptime_data(result1),
                    crash_data=result1.crash_data,
                    ok_count=result1.ok_count,
                    ng_count=result1.ng_count,
                    total_count=result1.total_count,
                    anr_count=result1.anr_count,
                    fatal_count=result1.fatal_count,
                    app_start_kill_data=result1.app_start_kill_data
                )
                averaged_result2 = AnalysisResult(
                    prefix=result2.prefix,
                    uptime_data=self._get_averaged_uptime_data(result2),
                    crash_data=result2.crash_data,
                    ok_count=result2.ok_count,
                    ng_count=result2.ng_count,
                    total_count=result2.total_count,
                    anr_count=result2.anr_count,
                    fatal_count=result2.fatal_count,
                    app_start_kill_data=result2.app_start_kill_data
                )
            
            # Populate comparison sheets (using averaged data)
            self._create_io_read_comparison_sheet(ws_io_read_comparison, averaged_result1, averaged_result2)
            self._create_io_write_comparison_sheet(ws_io_write_comparison, averaged_result1, averaged_result2)
            
            # Populate app start/kill analysis sheet
            self._create_app_start_kill_sheet(ws_app_analysis, result1, result2)
            
            # Populate testing app PSS comparison sheet
            # Pass dumpstate contents to avoid re-reading files
            self._create_testing_app_pss_sheet(ws_testing_app_pss, result1, result2)
            
            # Populate PSS analysis sheet
            # Pass dumpstate contents to avoid re-reading files
            self._create_pss_analysis_sheet(ws_pss_analysis, result1, result2)
            
            # Create compiler comparison sheet
            ws_compiler_comparison = wb.create_sheet("Compiler_Comparison")
            self._create_compiler_comparison_sheet(ws_compiler_comparison, result1, result2)
            
            # Save with error handling
            try:
                wb.save(output_path)
            except PermissionError:
                print(f"Permission denied saving {output_path}")
                # Try to save with a different name
                backup_path = output_path.with_name(
                    output_path.stem + "_backup" + output_path.suffix
                )
                try:
                    wb.save(backup_path)
                    print(f"Excel saved as backup: {backup_path}")
                except PermissionError:
                    # Try to save with a timestamp
                    import datetime
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    timestamp_path = output_path.with_name(
                        output_path.stem + f"_backup_{timestamp}" + output_path.suffix
                    )
                    wb.save(timestamp_path)
                    print(f"Excel saved with timestamp: {timestamp_path}")
            return True
            
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
    
    def _create_device_sheet(self, ws, result: AnalysisResult, device_type: str):
        """Create optimized device sheet with bulk operations"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        self._setup_sheet_headers(ws, result.prefix, device_type, styles)
        
        # Prepare data in bulk
        data_rows = []
        for item in sorted(result.uptime_data, key=lambda x: x.filename):
            raw_display = (item.raw_line[:self.config.MAX_RAW_LOG_LENGTH] + "..." 
                         if len(item.raw_line) > self.config.MAX_RAW_LOG_LENGTH 
                         else item.raw_line)
            
            data_rows.append([
                item.filename,
                item.uptime_minutes if item.uptime_minutes is not None else "N/A",
                item.status,
                raw_display
            ])
        
        # Write data in bulk
        for row_idx, row_data in enumerate(data_rows, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                
                # Apply styling
                if col_idx == 1:  # File name
                    cell.alignment = Alignment(horizontal='left')
                elif col_idx == 2:  # Uptime
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 3:  # Status
                    cell.alignment = Alignment(horizontal='center')
                    if value == self.config.STATUS_OK:
                        cell.fill = styles['ok_fill']
                    elif value == self.config.STATUS_NG:
                        cell.fill = styles['ng_fill']
                elif col_idx == 4:  # Raw log
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
        
        # Add summary section
        self._add_summary_section(ws, result, styles)
        
        # Set column widths
        for col, width in self.config.COLUMN_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_crash_sheet(self, ws, result: AnalysisResult, device_type: str):
        """Create crash analysis sheet with FATAL/ANR data"""
        # Setup styles
        styles = self._create_styles()
        
        # Create crash-specific styles
        anr_fill = PatternFill(start_color="FFE6B8", end_color="FFE6B8", fill_type="solid")  # Light orange
        fatal_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")  # Light red
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result.prefix} FATAL/ANR Crash Analysis ({device_type})")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        
        # Column headers
        headers = ["File Name", "Crash Type", "App Name", "PID", "Raw Log"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Prepare data in bulk
        data_rows = []
        for crash in sorted(result.crash_data, key=lambda x: (x.filename, x.crash_type, x.app_name)):
            raw_display = (crash.raw_line[:self.config.MAX_RAW_LOG_LENGTH] + "..." 
                         if len(crash.raw_line) > self.config.MAX_RAW_LOG_LENGTH 
                         else crash.raw_line)
            
            data_rows.append([
                crash.filename,
                crash.crash_type,
                crash.app_name,
                crash.pid if crash.pid else "N/A",
                raw_display
            ])
        
        # Write data in bulk
        for row_idx, row_data in enumerate(data_rows, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                
                # Apply styling
                if col_idx == 1:  # File name
                    cell.alignment = Alignment(horizontal='left')
                elif col_idx == 2:  # Crash type
                    cell.alignment = Alignment(horizontal='center')
                    if value == "ANR":
                        cell.fill = anr_fill
                    elif value == "FATAL":
                        cell.fill = fatal_fill
                elif col_idx == 3:  # App name
                    cell.alignment = Alignment(horizontal='left')
                elif col_idx == 4:  # PID
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 5:  # Raw log
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
        
        # Add crash summary section
        self._add_crash_summary_section(ws, result, styles)
        
        # Set column widths for crash sheet
        crash_column_widths = {
            1: 40,  # File name
            2: 15,  # Crash type
            3: 30,  # App name
            4: 15,  # PID
            5: 80   # Raw log
        }
        
        for col, width in crash_column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _add_crash_summary_section(self, ws, result: AnalysisResult, styles: Dict[str, Any]):
        """Add crash summary statistics section"""
        row = len(result.crash_data) + 5  # Start after data rows
        
        # Summary title
        summary_cell = ws.cell(row, 1, "CRASH SUMMARY")
        summary_cell.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        
        # Statistics
        stats = [
            ("Total Crashes:", len(result.crash_data)),
            ("ANR Count:", result.anr_count),
            ("FATAL Count:", result.fatal_count)
        ]
        
        for label, value in stats:
            ws.cell(row, 1, label).font = Font(bold=True)
            value_cell = ws.cell(row, 2, value)
            value_cell.font = Font(bold=True)
            
            # Color code crash types
            if "ANR" in label:
                value_cell.fill = PatternFill(start_color="FFE6B8", end_color="FFE6B8", fill_type="solid")
            elif "FATAL" in label:
                value_cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
            
            row += 1
        
        # Overall crash assessment
        row += 1
        if result.anr_count == 0 and result.fatal_count == 0:
            overall_result = "NO CRASHES DETECTED"
            result_fill = styles['ok_fill']
        else:
            overall_result = "CRASHES DETECTED"
            result_fill = styles['ng_fill']
        
        ws.cell(row, 1, "Overall Result:").font = Font(bold=True)
        result_cell = ws.cell(row, 2, overall_result)
        result_cell.fill = result_fill
        result_cell.font = Font(bold=True)
    
    def _create_io_read_sheet(self, ws, result: AnalysisResult, device_type: str):
        """Create I/O read sheet following the structure of IO READ report template.xlsx"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result.prefix} I/O Read Analysis ({device_type})")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        
        # Column headers
        headers = ["File Name", "Top IO READ Process", "Amount (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Prepare data in bulk
        row_idx = 3
        for item in sorted(result.uptime_data, key=lambda x: x.filename):
            # Get I/O read data for this file
            # Use the I/O read data stored in the UptimeData object
            io_read_data = item.io_read_data if item.io_read_data else []
            
            # If we have I/O read data, add it to the sheet
            # Limit to top 10 processes as requested
            if io_read_data:
                # Only show top 10 processes
                top_10_processes = io_read_data[:10]
                
                first_process, first_amount = top_10_processes[0] if top_10_processes else ("", 0.0)
                
                # First row with file name and top process
                ws.cell(row_idx, 1, item.filename)
                ws.cell(row_idx, 2, first_process)
                ws.cell(row_idx, 3, first_amount)
                row_idx += 1
                
                # Additional rows for other processes (without file name)
                for process, amount in top_10_processes[1:]:
                    ws.cell(row_idx, 1, "")  # Empty file name
                    ws.cell(row_idx, 2, process)
                    ws.cell(row_idx, 3, amount)
                    row_idx += 1
            else:
                # If no I/O read data, still add the file name
                ws.cell(row_idx, 1, item.filename)
                ws.cell(row_idx, 2, "No I/O read data")
                ws.cell(row_idx, 3, 0.0)
                row_idx += 1
        
        # Set column widths
        io_read_column_widths = {
            1: 40,  # File name
            2: 25,  # Top IO READ Process
            3: 15   # Amount (MB)
        }
        
        for col, width in io_read_column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_io_write_sheet(self, ws, result: AnalysisResult, device_type: str):
        """Create I/O write sheet following the structure of IO WRITE report template.xlsx"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result.prefix} I/O Write Analysis ({device_type})")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        
        # Column headers
        headers = ["File Name", "Top IO WRITE Process", "Amount (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Prepare data in bulk
        row_idx = 3
        for item in sorted(result.uptime_data, key=lambda x: x.filename):
            # Get I/O write data for this file
            # Use the I/O write data stored in the UptimeData object
            io_write_data = item.io_write_data if item.io_write_data else []
            
            # If we have I/O write data, add it to the sheet
            # Limit to top 10 processes as requested
            if io_write_data:
                # Only show top 10 processes
                top_10_processes = io_write_data[:10]
                
                first_process, first_amount = top_10_processes[0] if top_10_processes else ("", 0.0)
                
                # First row with file name and top process
                ws.cell(row_idx, 1, item.filename)
                ws.cell(row_idx, 2, first_process)
                ws.cell(row_idx, 3, first_amount)
                row_idx += 1
                
                # Additional rows for other processes (without file name)
                for process, amount in top_10_processes[1:]:
                    ws.cell(row_idx, 1, "")  # Empty file name
                    ws.cell(row_idx, 2, process)
                    ws.cell(row_idx, 3, amount)
                    row_idx += 1
            else:
                # If no I/O write data, still add the file name
                ws.cell(row_idx, 1, item.filename)
                ws.cell(row_idx, 2, "No I/O write data")
                ws.cell(row_idx, 3, 0.0)
                row_idx += 1
        
        # Set column widths
        io_write_column_widths = {
            1: 40,  # File name
            2: 25,  # Top IO WRITE Process
            3: 15   # Amount (MB)
        }
        
        for col, width in io_write_column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_styles(self) -> Dict[str, Any]:
        """Create reusable styles"""
        return {
            'header_font': Font(bold=True, color="FFFFFF"),
            'header_alignment': Alignment(horizontal='center'),
            'header_fill': PatternFill(
                start_color=self.config.COLORS['header'], 
                end_color=self.config.COLORS['header'], 
                fill_type="solid"
            ),
            'ok_fill': PatternFill(
                start_color=self.config.COLORS['ok'], 
                end_color=self.config.COLORS['ok'], 
                fill_type="solid"
            ),
            'ng_fill': PatternFill(
                start_color=self.config.COLORS['ng'], 
                end_color=self.config.COLORS['ng'], 
                fill_type="solid"
            )
        }
    
    def _setup_sheet_headers(self, ws, prefix: str, device_type: str, styles: Dict[str, Any]):
        """Setup sheet headers and styling"""
        # Title row
        title_cell = ws.cell(1, 1, f"{prefix} Uptime Status Analysis ({device_type})")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        # Column headers
        headers = ["File Name", "Uptime (minutes)", "Status", "Raw Log"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
    
    def _get_averaged_uptime_data(self, result: AnalysisResult) -> List[UptimeData]:
        """Get uptime data with averaged IO values by part"""
        # Group uptime data by parts and calculate averages for IO READ/WRITE
        part_io_data = self._group_and_calculate_averages(result.uptime_data)
        
        # Update uptime_data with averaged IO data
        updated_uptime_data = self._update_uptime_data_with_averages(result.uptime_data, part_io_data)
        
        return updated_uptime_data
    
    
    def _add_summary_section(self, ws, result: AnalysisResult, styles: Dict[str, Any]):
        """Add summary statistics section"""
        row = len(result.uptime_data) + 5  # Start after data rows
        
        # Summary title
        summary_cell = ws.cell(row, 1, "SUMMARY")
        summary_cell.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        # Statistics
        stats = [
            ("Total Files:", result.total_count),
            (f"OK Status (< {self.config.UPTIME_THRESHOLD_MINUTES} min):", result.ok_count),
            (f"NG Status (> {self.config.UPTIME_THRESHOLD_MINUTES} min):", result.ng_count)
        ]
        
        for label, value in stats:
            ws.cell(row, 1, label).font = Font(bold=True)
            value_cell = ws.cell(row, 2, value)
            value_cell.font = Font(bold=True)
            
            if "OK" in label:
                value_cell.fill = styles['ok_fill']
            elif "NG" in label:
                value_cell.fill = styles['ng_fill']
            
            row += 1
        
        # Overall result
        row += 1
        overall_result = ("ALL SYSTEMS GOOD" if result.ng_count == 0 
                         else "SYSTEMS HAVE ISSUES")
        result_fill = (styles['ok_fill'] if result.ng_count == 0 
                      else styles['ng_fill'])
        
        ws.cell(row, 1, "Overall Result:").font = Font(bold=True)
        result_cell = ws.cell(row, 2, overall_result)
        result_cell.fill = result_fill
        result_cell.font = Font(bold=True)
    
    def _create_io_read_comparison_sheet(self, ws, result1: AnalysisResult, result2: AnalysisResult):
        """Create I/O read comparison sheet showing DUT vs REF by part"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result1.prefix} vs {result2.prefix} I/O Read Comparison")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Column headers
        headers = ["Part", "DUT Process", "DUT Amount (MB)", "REF Process", "REF Amount (MB)", "Diff (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Group data by part for both DUT and REF
        dut_part_data = defaultdict(list)
        ref_part_data = defaultdict(list)
        
        # Collect DUT data by part
        for item in result1.uptime_data:
            if item.part_name and item.io_read_data:
                dut_part_data[item.part_name].extend(item.io_read_data)
        
        # Collect REF data by part
        for item in result2.uptime_data:
            if item.part_name and item.io_read_data:
                ref_part_data[item.part_name].extend(item.io_read_data)
        
        # Create comparison data by part
        row_idx = 3
        all_parts = set(dut_part_data.keys()) | set(ref_part_data.keys())
        
        for part in sorted(all_parts):
            # Get data for this part
            dut_processes = dut_part_data.get(part, [])
            ref_processes = ref_part_data.get(part, [])
            
            # Sort by amount descending and take top 10 for each
            dut_processes.sort(key=lambda x: x[1], reverse=True)
            dut_processes = dut_processes[:10]
            ref_processes.sort(key=lambda x: x[1], reverse=True)
            ref_processes = ref_processes[:10]
            
            # Create dictionaries for easier lookup
            dut_dict = {process: amount for process, amount in dut_processes}
            ref_dict = {process: amount for process, amount in ref_processes}
            
            # Get all unique processes for this part
            all_processes = set(dut_dict.keys()) | set(ref_dict.keys())
            
            # Create comparison data for this part
            part_comparison_data = []
            for process in all_processes:
                dut_amount = dut_dict.get(process, 0)
                ref_amount = ref_dict.get(process, 0)
                diff = dut_amount - ref_amount
                part_comparison_data.append((process, round(dut_amount), round(ref_amount), round(diff)))
            
            # Sort by maximum amount (descending)
            part_comparison_data.sort(key=lambda x: max(x[1], x[2]), reverse=True)
            
            # Write data to sheet
            for process, dut_amount, ref_amount, diff in part_comparison_data:
                ws.cell(row_idx, 1, part).alignment = Alignment(horizontal='center')  # Part
                ws.cell(row_idx, 2, process if dut_amount != 0 else "").alignment = Alignment(horizontal='center')  # DUT Process
                ws.cell(row_idx, 3, dut_amount if dut_amount != 0 else "").alignment = Alignment(horizontal='center')  # DUT Amount
                ws.cell(row_idx, 4, process if ref_amount != 0 else "").alignment = Alignment(horizontal='center')  # REF Process
                ws.cell(row_idx, 5, ref_amount if ref_amount != 0 else "").alignment = Alignment(horizontal='center')  # REF Amount
                diff_cell = ws.cell(row_idx, 6, diff)  # Diff
                diff_cell.alignment = Alignment(horizontal='center')
                # Highlight diff cells with diff value > IO threshold
                if abs(diff) > self.config.IO_DIFF_THRESHOLD:
                    diff_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
                row_idx += 1
        
        # Auto-fit column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-fit row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True
    
    def _extract_compiler_info(self, dumpstate_contents: Dict[Path, str]) -> Dict[str, str]:
        """
        Extract compiler information for test apps from dumpstate contents.
        Only process the first dumpstate file to maintain performance.
        
        Args:
            dumpstate_contents: Dictionary mapping file paths to their contents
            
        Returns:
            Dict mapping app names to their compiler types
        """
        compiler_data = {}
        
        # Test apps and their package names
        test_apps = {
            'camera': 'com.sec.android.app.camera',
            'helloworld': 'com.samsung.performance.helloworld_v6',
            'calllog': 'com.samsung.android.dialer',
            'dial': 'com.samsung.android.dialer',
            'clock': 'com.sec.android.app.clockpackage',
            'contact': 'com.samsung.android.app.contacts',
            'calendar': 'com.samsung.android.calendar',
            'calculator': 'com.samsung.android.calendar',
            'gallery': 'com.sec.android.gallery3d',
            'message': 'com.samsung.android.messaging',
            'menu': 'com.sec.android.app.launcher',
            'myfile': 'com.sec.android.app.myfiles',
            'sip': 'com.samsung.android.honeyboard',
            'internet': 'com.sec.android.app.sbrowser',
            'note': 'com.samsung.android.app.notes',
            'setting': 'com.android.settings',
            'voice': 'com.sec.android.app.voicenote',
            'recent': 'com.sec.android.app.launcher'
        }
        
        # Process only the first dumpstate content to maintain performance
        # Convert to list and take the first item
        dumpstate_items = list(dumpstate_contents.items())
        if not dumpstate_items:
            return compiler_data
            
        file_path, content = dumpstate_items[0]
        if not content:
            return compiler_data
                
        # Look for Dexopt state section
        dexopt_start = content.find("Dexopt state:")
        if dexopt_start == -1:
            return compiler_data
                
        # Look for Compiler stats section
        compiler_stats_start = content.find("Compiler stats:")
        if compiler_stats_start == -1:
            compiler_stats_start = len(content)
            
        # Extract the Dexopt state section
        dexopt_section = content[dexopt_start:compiler_stats_start]
        
        # For each test app, find its compiler information
        for app_name, package_name in test_apps.items():
            # Look for the package name in the Dexopt state section
            package_start = dexopt_section.find(f"[{package_name}]")
            if package_start == -1:
                continue
                
            # Look for arm64 line within 10 lines after the package name
            lines = dexopt_section[package_start:].split('\n')
            for i in range(min(10, len(lines))):
                line = lines[i]
                if 'arm64:' in line and '[status=' in line:
                    # Extract status from [status=XXX]
                    status_match = re.search(r'\[status=([^\]]+)\]', line)
                    if status_match:
                        compiler_status = status_match.group(1)
                        compiler_data[app_name] = compiler_status
                        break
        
        return compiler_data
    
    def _create_pss_analysis_sheet(self, ws, result1: AnalysisResult, result2: AnalysisResult):
        """Create PSS analysis sheet merged from separate PSS report"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result1.prefix} vs {result2.prefix} PSS Analysis")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        
        # Column headers
        headers = ["Folder Name", "Process Name", "PSS (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Import the get_ram_size function
        try:
            from .analyze_pss import get_ram_size
        except ImportError:
            from analyze_pss import get_ram_size
        
        # Determine RAM sizes and thresholds for DUT and REF
        dut_threshold = self.config.DEFAULT_PSS_THRESHOLD_MB  # Default threshold from Config
        ref_threshold = self.config.DEFAULT_PSS_THRESHOLD_MB  # Default threshold from Config
        
        # Get the first dumpstate file for DUT to determine RAM size
        # Use stored dumpstate contents if available
        for item in result1.uptime_data:
            if item.extracted_file_path:
                try:
                    if result1.dumpstate_contents and item.extracted_file_path in result1.dumpstate_contents:
                        content = result1.dumpstate_contents[item.extracted_file_path]
                    else:
                        with open(item.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                            content = f.read()
                    ram_size = get_ram_size(content)
                    dut_threshold = self.config.get_threshold_for_ram(ram_size)
                    print(f"DUT RAM size: {ram_size}GB, PSS threshold: {dut_threshold}MB")
                    break
                except Exception as e:
                    print(f"Error reading DUT dumpstate file for RAM detection: {e}")
                    break
        
        # Get the first dumpstate file for REF to determine RAM size
        # Use stored dumpstate contents if available
        for item in result2.uptime_data:
            if item.extracted_file_path:
                try:
                    if result2.dumpstate_contents and item.extracted_file_path in result2.dumpstate_contents:
                        content = result2.dumpstate_contents[item.extracted_file_path]
                    else:
                        with open(item.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                            content = f.read()
                    ram_size = get_ram_size(content)
                    ref_threshold = self.config.get_threshold_for_ram(ram_size)
                    print(f"REF RAM size: {ram_size}GB, PSS threshold: {ref_threshold}MB")
                    break
                except Exception as e:
                    print(f"Error reading REF dumpstate file for RAM detection: {e}")
                    break
        
        # Collect PSS data from both DUT and REF
        pss_data = []
        
        # Process DUT data
        # Use stored dumpstate contents if available
        for item in result1.uptime_data:
            if item.extracted_file_path:
                try:
                    if result1.dumpstate_contents and item.extracted_file_path in result1.dumpstate_contents:
                        content = result1.dumpstate_contents[item.extracted_file_path]
                    else:
                        with open(item.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                            content = f.read()
                    
                    # Find the PSS section
                    pss_start = content.find("Total PSS by process:")
                    pss_end = content.find("Total PSS by OOM adjustment:")
                    
                    if pss_start != -1 and pss_end != -1:
                        pss_section = content[pss_start:pss_end]
                        lines = pss_section.split('\n')
                        
                        for line in lines:
                            # Match lines with PSS data: "    XXX,XXXK: process_name (pid XXX) ..."
                            match = re.match(r'\s+(\d{1,3}(?:,\d{3})*)K:\s*([^\s\(]+)', line)
                            if match:
                                pss_value_str = match.group(1).replace(',', '')
                                pss_value_kb = int(pss_value_str)
                                pss_value_mb = pss_value_kb / 1024.0
                                process_name = match.group(2)
                                
                                # Only include processes with PSS > dut_threshold (based on RAM size)
                                if pss_value_mb > dut_threshold:
                                    folder_name = f"DUT_{item.filename}"
                                    pss_data.append((folder_name, process_name, pss_value_mb))
                                else:
                                    # Since PSS data is sorted from big to small, we can break early
                                    break
                except Exception as e:
                    print(f"Error processing PSS data for {item.extracted_file_path}: {e}")
        
        # Process REF data
        # Use stored dumpstate contents if available
        for item in result2.uptime_data:
            if item.extracted_file_path:
                try:
                    if result2.dumpstate_contents and item.extracted_file_path in result2.dumpstate_contents:
                        content = result2.dumpstate_contents[item.extracted_file_path]
                    else:
                        with open(item.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                            content = f.read()
                    
                    # Find the PSS section
                    pss_start = content.find("Total PSS by process:")
                    pss_end = content.find("Total PSS by OOM adjustment:")
                    
                    if pss_start != -1 and pss_end != -1:
                        pss_section = content[pss_start:pss_end]
                        lines = pss_section.split('\n')
                        
                        for line in lines:
                            # Match lines with PSS data: "    XXX,XXXK: process_name (pid XXX) ..."
                            match = re.match(r'\s+(\d{1,3}(?:,\d{3})*)K:\s*([^\s\(]+)', line)
                            if match:
                                pss_value_str = match.group(1).replace(',', '')
                                pss_value_kb = int(pss_value_str)
                                pss_value_mb = pss_value_kb / 1024.0
                                process_name = match.group(2)
                                
                                # Only include processes with PSS > ref_threshold (based on RAM size)
                                if pss_value_mb > ref_threshold:
                                    folder_name = f"REF_{item.filename}"
                                    pss_data.append((folder_name, process_name, pss_value_mb))
                                else:
                                    # Since PSS data is sorted from big to small, we can break early
                                    break
                except Exception as e:
                    print(f"Error processing PSS data for {item.extracted_file_path}: {e}")
        
        # Write data
        row_idx = 3
        for folder_name, process_name, pss_value_mb in pss_data:
            ws.cell(row=row_idx, column=1, value=folder_name)
            ws.cell(row=row_idx, column=2, value=process_name)
            ws.cell(row=row_idx, column=3, value=round(pss_value_mb, 2))
            row_idx += 1
        
        # Auto-adjust column widths
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-adjust row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True
    
    def _create_testing_app_pss_sheet(self, ws, result1: AnalysisResult, result2: AnalysisResult):
        """Create testing app PSS comparison sheet - calculate average PSS for each app before diff"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result1.prefix} vs {result2.prefix} Testing App PSS Comparison")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        
        # Column headers
        headers = ["Part", "App Name", "DUT PSS (MB)", "REF PSS (MB)", "Diff (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Create a mapping of folder name to UptimeData for both DUT and REF
        dut_folder_map = {item.filename: item for item in result1.uptime_data}
        ref_folder_map = {item.filename: item for item in result2.uptime_data}
        
        # Create part mappings based on the folder structure
        dut_part_map = {}
        ref_part_map = {}
        
        # Map folders to parts for DUT
        for item in result1.uptime_data:
            if item.filename:
                part_name = self._extract_part_name(item.filename)
                if part_name:
                    dut_part_map[item.filename] = part_name
                    
        # Map folders to parts for REF
        for item in result2.uptime_data:
            if item.filename:
                part_name = self._extract_part_name(item.filename)
                if part_name:
                    ref_part_map[item.filename] = part_name
        
        # Group folders by part
        dut_part_folders = defaultdict(list)
        ref_part_folders = defaultdict(list)
        
        # Group DUT folders by part
        for folder_name in dut_folder_map.keys():
            part_name = dut_part_map.get(folder_name)
            if part_name:
                dut_part_folders[part_name].append(folder_name)
        
        # Group REF folders by part
        for folder_name in ref_folder_map.keys():
            part_name = ref_part_map.get(folder_name)
            if part_name:
                ref_part_folders[part_name].append(folder_name)
        
        # Sort folder names within each part
        for part_name in dut_part_folders:
            dut_part_folders[part_name].sort()
        for part_name in ref_part_folders:
            ref_part_folders[part_name].sort()
        
        # Write data to sheet
        row_idx = 3
        # Get all unique parts
        all_parts = set(dut_part_folders.keys()) | set(ref_part_folders.keys())
        for part_name in sorted(all_parts):
            # Get all apps for this part (including calllog, dial, clock for 2part)
            target_apps = []
            for app, part in FOLDER_APP_PART_MAPPING.items():
                if f"{part}part" == part_name:
                    target_apps.append(app)
            
            # Special handling for 2part to include calllog, dial, clock
            if part_name == "2part":
                # Add additional apps for 2part if not already included
                additional_apps = ["calllog", "dial", "clock"]
                for app in additional_apps:
                    if app not in target_apps:
                        target_apps.append(app)
            
            # Process each app for this part
            for target_app in target_apps:
                # Get package name for the target app
                target_package = APP_PACKAGE_MAPPING.get(target_app)
                if not target_package:
                    continue
                
                # Get folder lists for this part
                dut_folders = dut_part_folders.get(part_name, [])
                ref_folders = ref_part_folders.get(part_name, [])
                
                # If both are empty, skip this app
                if len(dut_folders) == 0 and len(ref_folders) == 0:
                    continue
                
                # Extract PSS data for all folders of this app in this part
                dut_pss_values = []
                ref_pss_values = []
                
                # Process DUT folders
                # Use stored dumpstate contents if available
                for dut_folder in dut_folders:
                    dut_data = dut_folder_map.get(dut_folder)
                    if dut_data and dut_data.extracted_file_path:
                        try:
                            if result1.dumpstate_contents and dut_data.extracted_file_path in result1.dumpstate_contents:
                                dut_content = result1.dumpstate_contents[dut_data.extracted_file_path]
                            else:
                                with open(dut_data.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                                    dut_content = f.read()
                            dut_pss = extract_pss_for_package(dut_content, target_package)
                            if dut_pss > 0:
                                dut_pss_values.append(dut_pss)
                        except Exception as e:
                            print(f"Error extracting PSS for {target_package} in {dut_folder}: {e}")
                
                # Process REF folders
                # Use stored dumpstate contents if available
                for ref_folder in ref_folders:
                    ref_data = ref_folder_map.get(ref_folder)
                    if ref_data and ref_data.extracted_file_path:
                        try:
                            if result2.dumpstate_contents and ref_data.extracted_file_path in result2.dumpstate_contents:
                                ref_content = result2.dumpstate_contents[ref_data.extracted_file_path]
                            else:
                                with open(ref_data.extracted_file_path, "r", encoding="utf-8", errors="ignore") as f:
                                    ref_content = f.read()
                            ref_pss = extract_pss_for_package(ref_content, target_package)
                            if ref_pss > 0:
                                ref_pss_values.append(ref_pss)
                        except Exception as e:
                            print(f"Error extracting PSS for {target_package} in {ref_folder}: {e}")
                
                # Calculate averages
                dut_avg_pss = sum(dut_pss_values) / len(dut_pss_values) if dut_pss_values else 0.0
                ref_avg_pss = sum(ref_pss_values) / len(ref_pss_values) if ref_pss_values else 0.0
                
                # Calculate diff
                diff = dut_avg_pss - ref_avg_pss
                
                # Only add to report if diff exceeds PSS threshold
                if abs(diff) >= self.config.PSS_DIFF_THRESHOLD:
                    # Part name
                    ws.cell(row_idx, 1, part_name).alignment = Alignment(horizontal='center')
                    
                    # App name
                    ws.cell(row_idx, 2, target_app).alignment = Alignment(horizontal='center')
                    
                    # DUT PSS (average)
                    if dut_avg_pss > 0:
                        ws.cell(row_idx, 3, round(dut_avg_pss, 2)).alignment = Alignment(horizontal='center')
                    else:
                        ws.cell(row_idx, 3, "").alignment = Alignment(horizontal='center')
                    
                    # REF PSS (average)
                    if ref_avg_pss > 0:
                        ws.cell(row_idx, 4, round(ref_avg_pss, 2)).alignment = Alignment(horizontal='center')
                    else:
                        ws.cell(row_idx, 4, "").alignment = Alignment(horizontal='center')
                    
                    # Diff
                    diff_cell = ws.cell(row_idx, 5, round(diff, 2))
                    diff_cell.alignment = Alignment(horizontal='center')
                    # Highlight diff cells with diff value >= PSS threshold
                    if abs(diff) >= self.config.PSS_DIFF_THRESHOLD:
                        diff_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
                    
                    row_idx += 1
        
        # Auto-fit column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-fit row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True
    
    def _create_io_write_comparison_sheet(self, ws, result1: AnalysisResult, result2: AnalysisResult):
        """Create I/O write comparison sheet showing DUT vs REF by part"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result1.prefix} vs {result2.prefix} I/O Write Comparison")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Column headers
        headers = ["Part", "DUT Process", "DUT Amount (MB)", "REF Process", "REF Amount (MB)", "Diff (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Group data by part for both DUT and REF
        dut_part_data = defaultdict(list)
        ref_part_data = defaultdict(list)
        
        # Collect DUT data by part
        for item in result1.uptime_data:
            if item.part_name and item.io_write_data:
                dut_part_data[item.part_name].extend(item.io_write_data)
        
        # Collect REF data by part
        for item in result2.uptime_data:
            if item.part_name and item.io_write_data:
                ref_part_data[item.part_name].extend(item.io_write_data)
        
        # Create comparison data by part
        row_idx = 3
        all_parts = set(dut_part_data.keys()) | set(ref_part_data.keys())
        
        for part in sorted(all_parts):
            # Get data for this part
            dut_processes = dut_part_data.get(part, [])
            ref_processes = ref_part_data.get(part, [])
            
            # Sort by amount descending and take top 10 for each
            dut_processes.sort(key=lambda x: x[1], reverse=True)
            dut_processes = dut_processes[:10]
            ref_processes.sort(key=lambda x: x[1], reverse=True)
            ref_processes = ref_processes[:10]
            
            # Create dictionaries for easier lookup
            dut_dict = {process: amount for process, amount in dut_processes}
            ref_dict = {process: amount for process, amount in ref_processes}
            
            # Get all unique processes for this part
            all_processes = set(dut_dict.keys()) | set(ref_dict.keys())
            
            # Create comparison data for this part
            part_comparison_data = []
            for process in all_processes:
                dut_amount = dut_dict.get(process, 0)
                ref_amount = ref_dict.get(process, 0)
                diff = dut_amount - ref_amount
                part_comparison_data.append((process, round(dut_amount), round(ref_amount), round(diff)))
            
            # Sort by maximum amount (descending)
            part_comparison_data.sort(key=lambda x: max(x[1], x[2]), reverse=True)
            
            # Write data to sheet
            for process, dut_amount, ref_amount, diff in part_comparison_data:
                ws.cell(row_idx, 1, part).alignment = Alignment(horizontal='center')  # Part
                ws.cell(row_idx, 2, process if dut_amount != 0 else "").alignment = Alignment(horizontal='center')  # DUT Process
                ws.cell(row_idx, 3, dut_amount if dut_amount != 0 else "").alignment = Alignment(horizontal='center')  # DUT Amount
                ws.cell(row_idx, 4, process if ref_amount != 0 else "").alignment = Alignment(horizontal='center')  # REF Process
                ws.cell(row_idx, 5, ref_amount if ref_amount != 0 else "").alignment = Alignment(horizontal='center')  # REF Amount
                diff_cell = ws.cell(row_idx, 6, diff)  # Diff
                diff_cell.alignment = Alignment(horizontal='center')
                # Highlight diff cells with diff value > IO threshold
                if abs(diff) > self.config.IO_DIFF_THRESHOLD:
                    diff_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
                row_idx += 1
        
        # Auto-fit column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-fit row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True
    
    def _create_app_start_kill_sheet(self, ws, result1: AnalysisResult, result2: AnalysisResult):
        """Create app start/kill analysis sheet - show each part 3 times (for each folder) and group by app name"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result1.prefix} vs {result2.prefix} App Start/Kill Analysis")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        
        # Column headers
        headers = ["Part", "App Name", "DUT Start Count", "DUT Start Reasons", "DUT Kill Count", "DUT Kill Reasons", "REF Start Count", "REF Start Reasons", "REF Kill Count", "REF Kill Reasons"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Create a mapping of folder name to list of AppStartKillInfo for both DUT and REF
        # Group by folder name since each folder can have multiple apps
        from collections import defaultdict
        dut_folder_map = defaultdict(list)
        ref_folder_map = defaultdict(list)
        
        for app_info in result1.app_start_kill_data:
            if app_info.folder_name:
                dut_folder_map[app_info.folder_name].append(app_info)
                
        for app_info in result2.app_start_kill_data:
            if app_info.folder_name:
                ref_folder_map[app_info.folder_name].append(app_info)
        
        # Create part mappings based on the folder structure
        dut_part_map = {}
        ref_part_map = {}
        
        # Map folders to parts for DUT
        for folder_name in dut_folder_map.keys():
            part_name = self._extract_part_name(folder_name)
            if part_name:
                dut_part_map[folder_name] = part_name
                    
        # Map folders to parts for REF
        for folder_name in ref_folder_map.keys():
            part_name = self._extract_part_name(folder_name)
            if part_name:
                ref_part_map[folder_name] = part_name
        
        # Group folders by part
        dut_part_folders = defaultdict(list)
        ref_part_folders = defaultdict(list)
        
        # Group DUT folders by part
        for folder_name in dut_folder_map.keys():
            part_name = dut_part_map.get(folder_name)
            if part_name:
                dut_part_folders[part_name].append(folder_name)
        
        # Group REF folders by part
        for folder_name in ref_folder_map.keys():
            part_name = ref_part_map.get(folder_name)
            if part_name:
                ref_part_folders[part_name].append(folder_name)
        
        # Sort folder names within each part
        for part_name in dut_part_folders:
            dut_part_folders[part_name].sort()
        for part_name in ref_part_folders:
            ref_part_folders[part_name].sort()
        
        # Write data to sheet - show each part 3 times and group by app name
        row_idx = 3
        
        # Get all unique apps that should be analyzed according to FOLDER_APP_PART_MAPPING
        all_expected_apps = list(FOLDER_APP_PART_MAPPING.keys())
        # Add special apps for 2part
        additional_2part_apps = ["calllog", "dial", "clock"]
        for app in additional_2part_apps:
            if app not in all_expected_apps:
                all_expected_apps.append(app)
        
        # Group all folders by part and sort them
        all_parts = set(dut_part_folders.keys()) | set(ref_part_folders.keys())
        
        # For each app, show its data for all parts where it belongs
        for expected_app in sorted(all_expected_apps):
            # Find which part this app belongs to
            app_part = None
            for app, part in FOLDER_APP_PART_MAPPING.items():
                if app == expected_app:
                    app_part = f"{part}part"
                    break
            
            # Special handling for calllog, dial, clock - they belong to 2part
            if expected_app in ["calllog", "dial", "clock"]:
                app_part = "2part"
            
            if not app_part:
                continue
                
            # Get folders for this part
            dut_folders = dut_part_folders.get(app_part, [])
            ref_folders = ref_part_folders.get(app_part, [])
            
            # Show data for each folder (3 times per part)
            max_folders = max(len(dut_folders), len(ref_folders))
            
            # Process each folder/run
            for folder_idx in range(max_folders):
                # Get folder names for this index
                dut_folder = dut_folders[folder_idx] if folder_idx < len(dut_folders) else None
                ref_folder = ref_folders[folder_idx] if folder_idx < len(ref_folders) else None
                
                # Get app info for this specific folder
                dut_app_info = None
                if dut_folder:
                    for app_info in dut_folder_map.get(dut_folder, []):
                        if app_info.app_name == expected_app:
                            dut_app_info = app_info
                            break
                
                ref_app_info = None
                if ref_folder:
                    for app_info in ref_folder_map.get(ref_folder, []):
                        if app_info.app_name == expected_app:
                            ref_app_info = app_info
                            break
                
                # Get values for this app in this specific folder
                dut_start_count = dut_app_info.start_count if dut_app_info else 0
                dut_start_reasons = ", ".join(dut_app_info.start_reasons) if dut_app_info and dut_app_info.start_reasons else ""
                dut_kill_count = dut_app_info.kill_count if dut_app_info else 0
                dut_kill_reasons = ", ".join(dut_app_info.kill_reasons) if dut_app_info and dut_app_info.kill_reasons else ""
                
                ref_start_count = ref_app_info.start_count if ref_app_info else 0
                ref_start_reasons = ", ".join(ref_app_info.start_reasons) if ref_app_info and ref_app_info.start_reasons else ""
                ref_kill_count = ref_app_info.kill_count if ref_app_info else 0
                ref_kill_reasons = ", ".join(ref_app_info.kill_reasons) if ref_app_info and ref_app_info.kill_reasons else ""
                
                # Show the part name (without timestamp)
                part_display = app_part
                
                # Part name
                ws.cell(row_idx, 1, part_display).alignment = Alignment(horizontal='center')
                
                # App name
                ws.cell(row_idx, 2, expected_app).alignment = Alignment(horizontal='center')
                
                # DUT info
                ws.cell(row_idx, 3, dut_start_count).alignment = Alignment(horizontal='center')
                ws.cell(row_idx, 4, dut_start_reasons).alignment = Alignment(horizontal='center')
                ws.cell(row_idx, 5, dut_kill_count).alignment = Alignment(horizontal='center')
                ws.cell(row_idx, 6, dut_kill_reasons).alignment = Alignment(horizontal='center')
                
                # REF info
                ws.cell(row_idx, 7, ref_start_count).alignment = Alignment(horizontal='center')
                ws.cell(row_idx, 8, ref_start_reasons).alignment = Alignment(horizontal='center')
                ws.cell(row_idx, 9, ref_kill_count).alignment = Alignment(horizontal='center')
                ws.cell(row_idx, 10, ref_kill_reasons).alignment = Alignment(horizontal='center')
                
                row_idx += 1
        
        # Auto-fit column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-fit row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True
        
        # Auto-fit column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-fit row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True
    
    def _create_compiler_comparison_sheet(self, ws, result1: AnalysisResult, result2: AnalysisResult):
        """Create compiler comparison sheet showing DUT vs REF compiler information"""
        # Setup styles
        styles = self._create_styles()
        
        # Setup headers
        title_cell = ws.cell(1, 1, f"{result1.prefix} vs {result2.prefix} Compiler Comparison")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        # Column headers
        headers = ["App Name", "DUT Compiler Status", "REF Compiler Status", "Match"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = styles['header_font']
            cell.alignment = styles['header_alignment']
            cell.fill = styles['header_fill']
        
        # Get all unique app names from both results
        all_apps = set(result1.compiler_data.keys()) | set(result2.compiler_data.keys())
        
        # Write data
        row_idx = 3
        for app_name in sorted(all_apps):
            dut_status = result1.compiler_data.get(app_name, "N/A")
            ref_status = result2.compiler_data.get(app_name, "N/A")
            match = "YES" if dut_status == ref_status else "NO"
            
            # App name
            ws.cell(row_idx, 1, app_name).alignment = Alignment(horizontal='center')
            
            # DUT compiler status
            ws.cell(row_idx, 2, dut_status).alignment = Alignment(horizontal='center')
            
            # REF compiler status
            ws.cell(row_idx, 3, ref_status).alignment = Alignment(horizontal='center')
            
            # Match status
            match_cell = ws.cell(row_idx, 4, match)
            match_cell.alignment = Alignment(horizontal='center')
            if match == "NO":
                match_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
            
            row_idx += 1
        
        # Auto-fit column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Auto-fit row heights
        for row in range(1, row_idx):
            ws.row_dimensions[row].auto_size = True


def analyze_device_performance(dut: Device, ref: Device, extracted: bool = False) -> str:
    """Main function with optimized implementation using Device OOP structure"""
    config = dut.config if hasattr(dut, 'config') else Config()
    
    # Create comparator
    comparator = DeviceComparator(dut, ref, config)
    
    # Compare devices to set anr_fatal and uptime attributes
    comparator.compare()
    
    # Generate report
    path1 = dut.folder_path
    output_path = path1 / f"DevicePerformance_{dut.get_prefix()}_{ref.get_prefix()}.xlsx"
    success = comparator.generate_excel_report(output_path)
    
    if success:
        print(f"Excel created: {output_path}")
    
    # Generate console summary
    console_result = comparator.generate_console_report()
    print(console_result)
    
    return console_result


if __name__ == "__main__":
    if len(sys.argv) == 3:
        folder1 = sys.argv[1]
        folder2 = sys.argv[2]
        analyze_device_performance(folder1, folder2)
    else:
        print("Usage: python device_performance_analyzer.py <folder1> <folder2>")
        print("Example: python device_performance_analyzer.py dut_logs ref_logs")
