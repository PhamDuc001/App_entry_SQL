import os
import re
import json
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def get_ram_size(dumpstate_content):
    '''
    MemTotal:        7373720 kB
    :param dumpstate_content:
    :return:
        7373720 kB
    '''
    list_ram = [2, 4, 6, 8, 12, 16, 18, 24, 32, 48]
    PATTERN_RAM_SIZE = r"^MemTotal:(.+?)\n"
    ram_size = re.findall(PATTERN_RAM_SIZE, dumpstate_content, re.MULTILINE)
    if len(ram_size) > 0:
        try:
            ram_size = ram_size[0].lower().replace("kb", "").strip()
            ram_size = math.floor(int(ram_size)/1000/1000)
            for ram in list_ram:
                if (ram_size + 1) <= ram:
                    ram_size = ram
                    break
        except:
            ram_size = 0
    else:
        ram_size = 8 # Default
    return ram_size

def get_debug_level(dumpstate_content):
    '''
    Extract debug level from dumpstate content.
    [ro.boot.debug_level]: [0x4f4c] -> LOW
    [ro.boot.debug_level]: [0x494d] -> MID
    [ro.boot.debug_level]: [0x4948] -> HIGH
    
    :param dumpstate_content: Content of the dumpstate file
    :return: Debug level as string ("LOW", "MID", "HIGH", or "UNKNOWN")
    '''
    # Pattern to match the debug level line
    DEBUG_LEVEL_PATTERN = r"\[ro\.boot\.debug_level\]:\s*\[(0x[0-9a-fA-F]+)\]"
    
    # Mapping of hex values to debug levels
    debug_level_map = {
        "0x4f4c": "LOW",
        "0x494d": "MID",
        "0x4948": "HIGH"
    }
    
    # Find the debug level in the content
    match = re.search(DEBUG_LEVEL_PATTERN, dumpstate_content)
    if match:
        hex_value = match.group(1).lower()
        return debug_level_map.get(hex_value, "UNKNOWN")
    else:
        return "UNKNOWN"

def get_threshold_for_ram(ram_size_gb):
    """
    Determine PSS threshold based on device RAM size.
    
    Args:
        ram_size_gb (int): Device RAM size in GB
        
    Returns:
        int: Threshold in MB
    """
    if ram_size_gb < 6:
        return 500  # 500MB threshold for devices with less than 6GB RAM
    elif ram_size_gb <= 8:
        return 800  # 800MB threshold for devices with 6-8GB RAM
    else:
        return 1024  # 1GB threshold for devices with more than 8GB RAM

def extract_pss_above_threshold(dumpstate_file_path, threshold=350):
    """
    Extract PSS data for processes that exceed the threshold.
    Optimized to break early when PSS values fall below threshold since data is sorted.
    
    Args:
        dumpstate_file_path (str): Path to the dumpstate file
        threshold (int): Threshold in MB for PSS filtering (default: 350)
        
    Returns:
        tuple: (folder_name, list of (process_name, pss_value_mb) tuples)
    """
    try:
        # Read the dumpstate file
        with open(dumpstate_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Get folder name
        folder_name = os.path.basename(os.path.dirname(dumpstate_file_path))
        
        # Find the PSS section
        pss_start = content.find("Total PSS by process:")
        pss_end = content.find("Total PSS by OOM adjustment:")
        
        if pss_start == -1 or pss_end == -1:
            print(f"Could not find PSS section in {dumpstate_file_path}")
            return (folder_name, [])
        
        # Extract the PSS section
        pss_section = content[pss_start:pss_end]
        
        # Parse PSS data lines and collect processes above threshold
        lines = pss_section.split('\n')
        high_pss_processes = []
        
        for line in lines:
            # Match lines with PSS data: "    XXX,XXXK: process_name (pid XXX) ..."
            match = re.match(r'\s+(\d{1,3}(?:,\d{3})*)K:\s*([^\s\(]+)', line)
            if match:
                pss_value_str = match.group(1).replace(',', '')
                pss_value_kb = int(pss_value_str)
                pss_value_mb = pss_value_kb / 1024.0
                process_name = match.group(2)
                
                # Check if PSS value exceeds threshold
                if pss_value_mb > threshold:
                    high_pss_processes.append((process_name, pss_value_mb))
                else:
                    # Since PSS data is sorted from big to small, we can break early
                    # when we encounter a value below the threshold
                    break
        
        return (folder_name, high_pss_processes)
        
    except Exception as e:
        print(f"Error processing {dumpstate_file_path}: {e}")
        return (os.path.basename(os.path.dirname(dumpstate_file_path)), [])

def find_and_extract_pss_data(dut_directory, ref_directory, pss_output_file):
    """
    Recursively search for dumpstate files in both DUT and REF directory structures and extract PSS data.
    
    Args:
        dut_directory (str): DUT directory to search for dumpstate files
        ref_directory (str): REF directory to search for dumpstate files
        pss_output_file (str): Path to the output PSS file (without extension)
    """
    all_pss_data_threshold = []
    
    # Determine RAM sizes and thresholds for DUT and REF
    dut_threshold = 350  # Default threshold
    ref_threshold = 350  # Default threshold
    
    # Get the first dumpstate file for DUT to determine RAM size
    if dut_directory and os.path.exists(dut_directory):
        dut_ram_detected = False
        for dirpath, dirnames, filenames in os.walk(dut_directory):
            if dut_ram_detected:
                break
            for filename in filenames:
                if filename.startswith('dumpstate-') and filename.endswith('.txt'):
                    dumpstate_file_path = os.path.join(dirpath, filename)
                    try:
                        with open(dumpstate_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                        ram_size = get_ram_size(content)
                        dut_threshold = get_threshold_for_ram(ram_size)
                        print(f"DUT RAM size: {ram_size}GB, PSS threshold: {dut_threshold}MB")
                        dut_ram_detected = True
                        break
                    except Exception as e:
                        print(f"Error reading DUT dumpstate file for RAM detection: {e}")
                        break
            if dut_ram_detected:
                break
    
    # Get the first dumpstate file for REF to determine RAM size
    if ref_directory and os.path.exists(ref_directory):
        ref_ram_detected = False
        for dirpath, dirnames, filenames in os.walk(ref_directory):
            if ref_ram_detected:
                break
            for filename in filenames:
                if filename.startswith('dumpstate-') and filename.endswith('.txt'):
                    dumpstate_file_path = os.path.join(dirpath, filename)
                    try:
                        with open(dumpstate_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                        ram_size = get_ram_size(content)
                        ref_threshold = get_threshold_for_ram(ram_size)
                        print(f"REF RAM size: {ram_size}GB, PSS threshold: {ref_threshold}MB")
                        ref_ram_detected = True
                        break
                    except Exception as e:
                        print(f"Error reading REF dumpstate file for RAM detection: {e}")
                        break
            if ref_ram_detected:
                break
    
    # Process both DUT and REF directories
    for root_directory, prefix, threshold in [(dut_directory, "DUT", dut_threshold), (ref_directory, "REF", ref_threshold)]:
        if not root_directory or not os.path.exists(root_directory):
            continue
            
        # Walk through the directory structure
        for dirpath, dirnames, filenames in os.walk(root_directory):
            # Look for dumpstate files
            for filename in filenames:
                if filename.startswith('dumpstate-') and filename.endswith('.txt'):
                    dumpstate_file_path = os.path.join(dirpath, filename)
                    folder_name = os.path.basename(dirpath)  # Get the bugreport folder name
                    
                    print(f"Processing: {dumpstate_file_path}")
                    
                    # Extract PSS data above threshold
                    folder_name, high_pss_processes = extract_pss_above_threshold(dumpstate_file_path, threshold)
                    if high_pss_processes:
                        # Add prefix to distinguish between DUT and REF
                        prefixed_folder_name = f"{prefix}_{folder_name}"
                        all_pss_data_threshold.append((prefixed_folder_name, high_pss_processes))
                
    # Create Excel report
    excel_output_file = pss_output_file.replace('.txt', '.xlsx') if pss_output_file.endswith('.txt') else pss_output_file
    # Ensure the file has .xlsx extension
    if not excel_output_file.endswith('.xlsx'):
        excel_output_file += '.xlsx'
    create_pss_excel_report(all_pss_data_threshold, excel_output_file)
    
    print(f"\nTotal {len(all_pss_data_threshold)} bugreport folders with processes above their respective thresholds")
    print(f"Results written to {excel_output_file}")

def analyze_pss_in_dumpstate_file(dumpstate_file_path, threshold=350):
    """
    Analyze PSS data in a single dumpstate file and return formatted results.
    
    Args:
        dumpstate_file_path (str): Path to the dumpstate file
        threshold (int): Threshold in MB for PSS filtering (default: 350)
        
    Returns:
        str: Formatted analysis results for inclusion in the report
    """
    folder_name, high_pss_processes = extract_pss_above_threshold(dumpstate_file_path, threshold)
    
    if not high_pss_processes:
        return ""
    
    result = f"\n\t- PSS Analysis (processes > {threshold}MB):\n"
    result += f"\t  {folder_name}:\n"
    
    for process_name, pss_mb in high_pss_processes:
        result += f"\t    {process_name}: {pss_mb:.1f} MB\n"
    
    return result

def create_pss_excel_report(pss_data, excel_output_file):
    """
    Create an Excel report from PSS data.
    
    Args:
        pss_data (list): List of (folder_name, high_pss_processes) tuples
        excel_output_file (str): Path to the output Excel file
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PSS Analysis"
    
    # Create styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # Write headers
    headers = ["Folder Name", "Process Name", "PSS (MB)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    row_idx = 2
    for folder_name, high_pss_processes in pss_data:
        for process_name, pss_value_mb in high_pss_processes:
            ws.cell(row=row_idx, column=1, value=folder_name)
            ws.cell(row=row_idx, column=2, value=process_name)
            ws.cell(row=row_idx, column=3, value=round(pss_value_mb, 2))
            row_idx += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(excel_output_file)
    print(f"Excel report saved to {excel_output_file}")

def extract_pss_for_package(dumpstate_content, package_name):
    """
    Extract PSS data for a specific package from dumpstate content.
    
    Args:
        dumpstate_content (str): Content of the dumpstate file
        package_name (str): Package name to extract PSS for
        
    Returns:
        float: PSS value in MB, or 0 if not found
    """
    try:
        # Find the PSS section
        pss_start = dumpstate_content.find("Total PSS by process:")
        pss_end = dumpstate_content.find("Total PSS by OOM adjustment:")
        
        if pss_start == -1 or pss_end == -1:
            return 0.0
        
        # Extract the PSS section
        pss_section = dumpstate_content[pss_start:pss_end]
        
        # Parse PSS data lines to find the specific package
        lines = pss_section.split('\n')
        for line in lines:
            # Match lines with PSS data: "    XXX,XXXK: process_name (pid XXX) ..."
            match = re.match(r'\s+(\d{1,3}(?:,\d{3})*)K:\s*([^\s\(]+)', line)
            if match:
                pss_value_str = match.group(1).replace(',', '')
                pss_value_kb = int(pss_value_str)
                pss_value_mb = pss_value_kb / 1024.0
                process_name = match.group(2)
                
                # Check if this is the package we're looking for
                if process_name == package_name:
                    return pss_value_mb
                    
        return 0.0
        
    except Exception as e:
        print(f"Error extracting PSS for package {package_name}: {e}")
        return 0.0

def analyze_pss_in_dumpstate_content(dumpstate_content, folder_name, threshold=350):
    """
    Analyze PSS data in dumpstate content and return formatted results.
    
    Args:
        dumpstate_content (str): Content of the dumpstate file
        folder_name (str): Name of the folder containing the dumpstate file
        threshold (int): Threshold in MB for PSS filtering (default: 350)
        
    Returns:
        str: Formatted analysis results for inclusion in the report
    """
    try:
        # Find the PSS section
        pss_start = dumpstate_content.find("Total PSS by process:")
        pss_end = dumpstate_content.find("Total PSS by OOM adjustment:")
        
        if pss_start == -1 or pss_end == -1:
            return ""
        
        # Extract the PSS section
        pss_section = dumpstate_content[pss_start:pss_end]
        
        # Parse PSS data lines and collect processes above threshold
        lines = pss_section.split('\n')
        high_pss_processes = []
        
        for line in lines:
            # Match lines with PSS data: "    XXX,XXXK: process_name (pid XXX) ..."
            match = re.match(r'\s+(\d{1,3}(?:,\d{3})*)K:\s*([^\s\(]+)', line)
            if match:
                pss_value_str = match.group(1).replace(',', '')
                pss_value_kb = int(pss_value_str)
                pss_value_mb = pss_value_kb / 1024.0
                process_name = match.group(2)
                
                # Check if PSS value exceeds threshold
                if pss_value_mb > threshold:
                    high_pss_processes.append((process_name, pss_value_mb))
                else:
                    # Since PSS data is sorted from big to small, we can break early
                    # when we encounter a value below the threshold
                    break
        
        if not high_pss_processes:
            return ""
        
        result = f"\n\t- PSS Analysis (processes > {threshold}MB):\n"
        result += f"\t  {folder_name}:\n"
        
        for process_name, pss_mb in high_pss_processes:
            result += f"\t    {process_name}: {pss_mb:.1f} MB\n"
        
        return result
        
    except Exception as e:
        print(f"Error processing PSS data: {e}")
        return ""

def main():
    """
    Main function to demonstrate PSS extraction.
    """
    # Example usage
    dut_directory = "test/full_test_dump"
    ref_directory = "test/full_test_dump"
    pss_output_file = "test/PSS_report"
    
    print("Extracting PSS data from dumpstate files...")
    print("=" * 50)
    
    find_and_extract_pss_data(dut_directory, ref_directory, pss_output_file)
    
    print("\nDone!")

if __name__ == "__main__":
    main()
