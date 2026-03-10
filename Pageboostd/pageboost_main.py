import os
import re
import sys
import zipfile
import shutil
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from functools import lru_cache

# ============================================================
# OPTIMIZATION: Pre-compile regex patterns at module level
# ============================================================
PAGEBOOSTD_PATTERN = re.compile(r"app\s+(\S+)\s+data_amount\s+(\d+)")

# ============================================================
# OPTIMIZATION: Cache for parsed pageboostd data
# ============================================================
_pageboostd_cache = {}

def clear_pageboostd_cache():
    """Clear the pageboostd cache"""
    global _pageboostd_cache
    _pageboostd_cache.clear()

def parse_pageboostd_cached(file_path):
    """
    Parse pageboostd file with caching.
    Returns dict with {app: value}
    """
    # Check cache first
    if file_path in _pageboostd_cache:
        return _pageboostd_cache[file_path]
    
    results = {}
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                m = PAGEBOOSTD_PATTERN.search(line)
                if m:
                    app = m.group(1)
                    val = int(m.group(2))
                    results[app] = val
    except Exception as e:
        print(f"[WARN] Cannot parse pageboostd file {file_path}: {e}")
    
    # Cache the result
    _pageboostd_cache[file_path] = results
    return results

def parse_pageboostd(file_path):
    """Backward-compatible wrapper for cached version"""
    return parse_pageboostd_cached(file_path)


APP_MAPPING = {
    "comsamsungperformancehelloworld_v6": "Helloworld",
    "comsamsungandroiddialer": "Dial",
    "comsecandroidappclockpackage": "Clock",
    "comsecandroidappcamera": "Camera",
    "comsamsungandroidappcontacts": "Contacts",
    "comsamsungandroidcalendar": "Calendar",
    "comsecandroidapppopupcalculator": "Calculator",
    "comsecandroidgallery3d": "Gallery",
    "comsamsungandroidmessaging": "Messages",
    "comsecandroidappmyfiles": "MyFiles",
    "comexampleedittexttest3": "SIP",
    "comsecandroidappsbrowser": "Internet",
    "comsamsungandroidappnotes": "Notes",
    "comandroidsettings": "Settings",
    "comsecandroidappvoicenote": "VoiceNote",
    "comgoogleandroidappsmessaging": "Messages",
}


def parse_pageboostd_from_zip(zip_path):
    """
    OPTIMIZATION: Stream ZIP content directly without extracting to disk.
    Returns parsed pageboostd data {app: value}
    """
    try:
        with zipfile.ZipFile(zip_path, 'r') as z:
            infos = z.infolist()
            if not infos:
                return {}
            
            # Find largest .txt file
            largest = None
            max_size = 0
            for info in infos:
                if not info.is_dir() and info.filename.lower().endswith('.txt'):
                    if info.file_size > max_size:
                        max_size = info.file_size
                        largest = info
            
            if not largest:
                return {}
            
            # Read and parse directly from ZIP
            with z.open(largest) as f:
                content = f.read().decode('utf-8', errors='ignore')
                results = {}
                for line in content.split('\n'):
                    m = PAGEBOOSTD_PATTERN.search(line)
                    if m:
                        app = m.group(1)
                        val = int(m.group(2))
                        results[app] = val
                return results
                
    except Exception as e:
        print(f"[WARN] Cannot parse ZIP {zip_path}: {e}")
        return {}


def collect_cycles_from_zips(folder):
    """
    OPTIMIZATION: Collect cycles by parsing ZIP files directly without extraction.
    """
    cycles = []  # list[dict], mỗi dict: {app: value}
    
    for fname in sorted(os.listdir(folder)):
        if not fname.lower().endswith(".zip"):
            continue
        fpath = os.path.join(folder, fname)
        if not os.path.isfile(fpath):
            continue

        # OPTIMIZATION: Parse directly from ZIP without extraction
        data = parse_pageboostd_from_zip(fpath)

        for app, val in data.items():
            placed = False
            for cyc in cycles:
                if app not in cyc:
                    cyc[app] = val
                    placed = True
                    break
            if not placed:
                cycles.append({app: val})
    #print(cycles)
    return cycles


def collect_cycles_from_extracted(folder):
    # Find largest file in each folder
    cycles = []
    for sub in sorted(os.listdir(folder)):
        subpath = os.path.join(folder, sub)
        if not os.path.isdir(subpath):
            continue

        # Find largest file in the folder
        largest_file = None
        max_size = -1
        for root, _, files in os.walk(subpath):
            for f in files:
                fpath = os.path.join(root, f)
                try:
                    size = os.path.getsize(fpath)
                except OSError:
                    continue
                if size > max_size:
                    max_size = size
                    largest_file = fpath
        if not largest_file:
            continue

        data = parse_pageboostd(largest_file)
        for app, val in data.items():
            placed = False
            for cyc in cycles:
                if app not in cyc:
                    cyc[app] = val
                    placed = True
                    break
            if not placed:
                cycles.append({app: val})
    #print(cycles)
    return cycles




def get_prefix(folder):
    # If .zip option
    for fname in os.listdir(folder):
        if fname.lower().endswith(".zip"):
            tokens = re.split(r"[_\-]", fname)
            if len(tokens) >= 2:
                return f"{tokens[0]}-{tokens[1]}-{tokens[2]}"
    # If folder extracted option
    for sub in os.listdir(folder):
        subpath = os.path.join(folder, sub)
        if os.path.isdir(subpath):
            tokens = re.split(r"[_\-]", sub)
            if len(tokens) >= 2:
                return f"{tokens[0]}-{tokens[1]}-{tokens[2]}"
    return "UNKNOWN"


def write_excel(out_path, prefix1, prefix2, cycles1, cycles2):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pageboostd"

    # style
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    # Header row 1: prefix1, prefix2
    n1, n2 = len(cycles1), len(cycles2)
    ws.cell(1, 1).value = "Unit: MB"
    ws.cell(1, 1).font = Font(bold=True)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    ws.cell(1, 2).value = prefix1
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+n1+1)
    ws.cell(1, 2).alignment = Alignment(horizontal="center")
    ws.cell(1, 2).font = Font(bold=True)

    ws.cell(1, 2+n1+1).value = prefix2
    ws.merge_cells(start_row=1, start_column=2+n1+1, end_row=1, end_column=1+n1+n2+2)
    ws.cell(1, 2+n1+1).alignment = Alignment(horizontal="center")
    ws.cell(1, 2+n1+1).font = Font(bold=True)

    # Header row 2: cycles
    for i in range(n1):
        ws.cell(2, 2+i).value = f"Cycle{i+1}"
        ws.cell(2, 2+i).font = header_font
        ws.cell(2, 2+i).alignment = header_alignment
    ws.cell(2, 2+n1).value = "Avg"
    ws.cell(2, 2+n1).font = header_font
    ws.cell(2, 2+n1).alignment = header_alignment

    for i in range(n2):
        ws.cell(2, 2+n1+1+i).value = f"Cycle{i+1}"
        ws.cell(2, 2+n1+1+i).font = header_font
        ws.cell(2, 2+n1+1+i).alignment = header_alignment
    ws.cell(2, 2+n1+1+n2).value = "Avg"
    ws.cell(2, 2+n1+1+n2).font = header_font
    ws.cell(2, 2+n1+1+n2).alignment = header_alignment

    ws.cell(2, 2+n1+1+n2+1).value = "Diff"
    ws.cell(2, 2+n1+1+n2+1).font = header_font
    ws.cell(2, 2+n1+1+n2+1).alignment = header_alignment

    apps = []
    seen = set()
    for cyc in cycles1 + cycles2:
        for app in cyc:
            if app not in seen:
                apps.append(app)
                seen.add(app)

    # Fill data
    row = 3
    for app in apps:
        label = APP_MAPPING.get(app, app)
        ws.cell(row, 1).value = label
        ws.cell(row, 1).alignment = header_alignment

        vals1 = [cyc.get(app) for cyc in cycles1]
        vals2 = [cyc.get(app) for cyc in cycles2]

        # folder1 cycles
        for i, v in enumerate(vals1):
            ws.cell(row, 2+i).value = round(v / 1000000, 2) if v else None
            ws.cell(row, 2+i).alignment = header_alignment
        avg1 = sum(v for v in vals1 if v is not None) / max(1, sum(1 for v in vals1 if v is not None))
        ws.cell(row, 2+n1).value = round(avg1 / 1000000, 2) if vals1 else None
        ws.cell(row, 2+n1).alignment = header_alignment

        # folder2 cycles
        for i, v in enumerate(vals2):
            ws.cell(row, 2+n1+1+i).value = round(v / 1000000, 2) if v else None
            ws.cell(row, 2+n1+1+i).alignment = header_alignment
        avg2 = sum(v for v in vals2 if v is not None) / max(1, sum(1 for v in vals2 if v is not None))
        ws.cell(row, 2+n1+1+n2).value = round(avg2 / 1000000, 2) if vals2 else None
        ws.cell(row, 2+n1+1+n2).alignment = header_alignment

        # diff
        if vals1 and vals2 and avg1 and avg2:
            ws.cell(row, 2+n1+1+n2+1).value = round((avg1 - avg2) / 1000000, 2)
            ws.cell(row, 2+n1+1+n2+1).alignment = header_alignment
        row += 1

    # Autofit
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(out_path)


def diff_pageboostd(folder1, folder2, extracted=False):
    # print("[OPTIMIZATION] Pageboostd mode - streaming ZIP content, no disk extraction")  # COMMENTED: Reduce log noise
    
    prefix1 = get_prefix(folder1)
    prefix2 = get_prefix(folder2)

    if extracted:
        cycles1 = collect_cycles_from_extracted(folder1)
        cycles2 = collect_cycles_from_extracted(folder2)
    else:
        cycles1 = collect_cycles_from_zips(folder1)
        cycles2 = collect_cycles_from_zips(folder2)

    out_path = os.path.join(folder1, f"ComparePageboostd_{prefix1}_{prefix2}.xlsx")
    write_excel(out_path, prefix1, prefix2, cycles1, cycles2)
    print(f"Excel created: {out_path}")
    
    # OPTIMIZATION: No longer needs cleanup since we don't extract to disk





def main():
    if len(sys.argv) < 3:
        print("Usage: python main.py <folder_dut> <folder_ref> [--extracted]")
        sys.exit(1)

    folder1, folder2 = sys.argv[1], sys.argv[2]
    extracted = len(sys.argv) > 3 and sys.argv[3] == "--extracted"

    diff_pageboostd(folder1, folder2, extracted)


if __name__ == "__main__":
    main()
